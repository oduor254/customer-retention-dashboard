from flask import Flask, render_template, jsonify, request
import pandas as pd
import numpy as np
from datetime import datetime
import json
import time
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
import requests
import io
import os
import threading
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures

from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__)

# Custom JSON encoder so jsonify() handles numpy scalar types without crashing.
from flask.json.provider import DefaultJSONProvider as _DefaultJSONProvider

class _NumpyJSONProvider(_DefaultJSONProvider):
    def default(self, obj):
        if hasattr(obj, 'item'):          # numpy scalars (int64, float64, bool_)
            return obj.item()
        if hasattr(obj, 'tolist'):        # numpy arrays
            return obj.tolist()
        return super().default(obj)

app.json_provider_class = _NumpyJSONProvider
app.json = _NumpyJSONProvider(app)

# Configuration
# Optional: Configure the start year for analysis. If not set, include all history.
DATA_START_YEAR_ENV = os.environ.get("DATA_START_YEAR")
try:
    DATA_START_YEAR = int(DATA_START_YEAR_ENV) if DATA_START_YEAR_ENV else None
except ValueError:
    DATA_START_YEAR = None

# Currency conversion rate (USD to KES)
USD_TO_KES = 1/130

# Color mapping by region
REGION_COLORS = {
    'Nairobi Metropolitan': '#3498db',
    'Nairobi CBD': "#1eda66ac",  # Blue
    'Coastal Region': '#e74c3c',        # Red
    'Western & Nyanza': '#2ecc71',      # Green
    'Central Region': '#f39c12',        # Orange
    'Rift Valley': '#9b59b6',           # Purple
    'Diaspora': '#1abc9c',
    'Website': '#27C2F5',
    'Rejects': '#27C2F5',               # Turquoise
}

# Shop to Region mapping
SHOP_REGION_MAP = {
    'Hazina': 'Nairobi CBD',
    'Hilton': 'Nairobi CBD',
    'Starmall': 'Nairobi CBD',
    'Ktda': 'Nairobi CBD',
    'Kitengela': 'Nairobi Metropolitan',
    'Rongai': 'Nairobi Metropolitan',
    'Rejects': 'Rejects',
    'Mombasa': 'Coastal Region',
    'Kakamega': 'Western & Nyanza',
    'Kisumu': 'Western & Nyanza',
    'Kisii': 'Western & Nyanza',
    'Busia': 'Western & Nyanza',
    'Meru': 'Central Region',
    'Nanyuki': 'Central Region',
    'Thika': 'Central Region',
    'Eldoret': 'Rift Valley',
    'Nakuru': 'Rift Valley',
    'Sinza': 'Diaspora',
    'Tanzania': 'Diaspora',
    'Uganda': 'Diaspora',
    'Website': 'Online',
}

# Cache variables
cached_data = None
last_fetch_time = None
computed_results_cache = None   # Legacy: full result (kept for compatibility)
global_results_cache  = None   # Global-only metrics (no shops) — fast path
shops_results_cache   = None   # Shop-by-shop metrics — slow path, loaded separately
CACHE_DURATION = 1800  # Cache for 30 minutes to reduce slow network calls
CACHE_FILE = '/tmp/customer_data_cache.csv' if os.name != 'nt' else 'customer_data_cache.csv'

# ── Supabase configuration ─────────────────────────────────────────────────
SUPABASE_URL  = os.environ.get("SUPABASE_URL", "https://nzxtvjulbucqcijqgive.supabase.co")
SUPABASE_KEY  = os.environ.get("SUPABASE_KEY")
DATABASE_URL  = os.environ.get("DATABASE_URL")   # postgres://... direct connection for fast bulk reads


# Column mapping: Supabase (snake_case) ↔ app (original names)
_SB_TO_APP = {
    'date': 'Date', 'first_name': 'First Name', 'gender': 'Gender',
    'phone': 'Phone', 'product': 'Product', 'color': 'Color',
    'category': 'Category', 'shop': 'Shop', 'price': 'Price',
    'quantity': 'Quantity', 'total': 'Total', 'month': 'Month',
    'month_year': 'Month-Year', 'quarter': 'Quarter',
    'marketing_expense': 'MARKETING EXPENSE',
}
_APP_TO_SB = {v: k for k, v in _SB_TO_APP.items()}


def _sb_headers():
    """Headers for direct PostgREST requests — no supabase-py client needed."""
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }


_SALES_COLS = ("date,first_name,gender,phone,product,color,category,shop,"
               "price,quantity,total,month,month_year,quarter,marketing_expense")


def _customer_id(phone_series):
    """Normalise a phone column into the customer key.

    Phone survives a CSV round-trip as a float ('794761620' -> 794761620.0), so
    the same customer read from the local working copy and from Supabase would
    otherwise produce two different keys and be counted as two people. Repeat
    -customer rates are the whole point of this dashboard, so the key has to be
    stable no matter which source the rows came from.
    """
    return (phone_series.astype(str).str.strip()
            .str.replace(r'\.0$', '', regex=True))

# Shared psycopg2 connect options.
#
# keepalives are the important part: when Supavisor drops a connection the
# server forgets the session entirely (pg_stat_activity shows nothing) while the
# client stays blocked on a socket that will never deliver another byte. No
# statement_timeout can save you there — the server is not running anything.
# Without these, a dropped read hangs the request forever; with them the OS
# tears the socket down and psycopg2 raises so the retry below can take over.
_PG_KWARGS = dict(
    connect_timeout=15,
    keepalives=1,
    keepalives_idle=30,
    keepalives_interval=10,
    keepalives_count=3,
    options='-c statement_timeout=120000',
)


def _load_from_supabase():
    """Fetch all rows via direct PostgreSQL connection — single query, ~2 seconds.

    Read in keyset-paginated chunks. Supavisor will not carry a quarter-million
    rows in a single response: a full-table SELECT dies with 'server closed the
    connection unexpectedly' / 'SSL error: unexpected eof', and COPY TO STDOUT
    fails even more reliably (3/3 on the transaction pooler, 2/2 on the session
    pooler). Short chunked queries stay well inside what the pooler tolerates.
    COPY *into* the table is fine — see _push_to_supabase_sql — only the
    outbound direction breaks.

    autocommit is essential: a read that dies mid-flight inside an open
    transaction leaves the backend 'idle in transaction' pinning an ACCESS SHARE
    lock on `sales`, which blocks the next sync's TRUNCATE indefinitely.
    """
    import psycopg2 # type: ignore

    # Measured on the live project: 4k-8k row chunks come back in 2-6s, while
    # anything larger regularly dies with 'SSL error: unexpected eof' or
    # 'server closed the connection unexpectedly'. Timings are erratic rather
    # than cleanly size-bound, so keep chunks small and lean on the retries.
    CHUNK = int(os.environ.get('SUPABASE_READ_CHUNK', 5_000))
    ATTEMPTS = 5
    frames, last_id, cols = [], 0, None

    while True:
        chunk_rows = None
        for attempt in range(1, ATTEMPTS + 1):
            conn = None
            try:
                conn = psycopg2.connect(DATABASE_URL, **_PG_KWARGS)
                conn.autocommit = True
                with conn.cursor() as cur:
                    cur.execute(
                        f"SELECT id,{_SALES_COLS} FROM sales "
                        "WHERE id > %s ORDER BY id LIMIT %s",
                        (last_id, CHUNK),
                    )
                    chunk_rows = cur.fetchall()
                    cols = [d[0] for d in cur.description]
                break
            except (psycopg2.OperationalError, psycopg2.DatabaseError) as e:
                if attempt == ATTEMPTS:
                    raise
                print(f"[WARNING] Supabase read chunk @id>{last_id} attempt "
                      f"{attempt}/{ATTEMPTS} failed ({type(e).__name__}: "
                      f"{str(e).strip()[:80]}); retrying...")
                time.sleep(2 * attempt)
            finally:
                if conn is not None:
                    conn.close()

        if not chunk_rows:
            break
        frames.append(pd.DataFrame(chunk_rows, columns=cols))
        last_id = chunk_rows[-1][0]
        if len(chunk_rows) < CHUNK:
            break

    if not frames:
        raise ValueError("Supabase sales table is empty — run a sync first.")

    df = pd.concat(frames, ignore_index=True)
    df.drop(columns=['id'], inplace=True)

    if df.empty:
        raise ValueError("Supabase sales table is empty — run a sync first.")

    df.rename(columns=_SB_TO_APP, inplace=True)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    for col in ('Price', 'Quantity', 'Total', 'MARKETING EXPENSE'):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    if 'Quantity' in df.columns:
        df['Quantity'] = df['Quantity'].replace(0, 1)
    df['Customer_ID'] = _customer_id(df['Phone'])
    print(f"[INFO] Loaded {len(df)} records from Supabase (direct SQL)")
    return df


def _ensure_supabase_tables():
    """Create the supporting tables if they are missing.

    Done in code rather than by hand so a fresh Supabase project works without
    anyone remembering to paste SQL into the dashboard.
    """
    import psycopg2 # type: ignore
    conn = psycopg2.connect(DATABASE_URL, **_PG_KWARGS)
    try:
        conn.autocommit = True
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS analytics_cache (
                    id         INTEGER PRIMARY KEY DEFAULT 1,
                    result     JSONB NOT NULL,
                    updated_at TIMESTAMPTZ DEFAULT NOW()
                )""")
    finally:
        conn.close()


def _push_to_supabase_sql(df_sb, truncate=True):
    """Write rows to the sales table over a direct PostgreSQL connection.

    truncate=True replaces the table (TRUNCATE + COPY); truncate=False appends
    (COPY only), which is what an incremental sync needs.

    The REST path could not do this: DELETE over ~250k rows blows past
    PostgREST's statement timeout (SQLSTATE 57014), and inserting via HTTP
    needs ~500 round-trips. TRUNCATE is O(1) regardless of row count and
    COPY streams every row in a single pass.

    `df_sb` must already use Supabase (snake_case) column names.
    """
    import csv as _csv
    import io as _io
    import psycopg2  # type: ignore

    cols = list(df_sb.columns)
    buf = _io.StringIO()
    writer = _csv.writer(buf, lineterminator='\n')
    for row in df_sb.itertuples(index=False, name=None):
        writer.writerow(
            ['' if v is None or (isinstance(v, float) and v != v) else v
             for v in row]
        )
    buf.seek(0)

    conn = psycopg2.connect(DATABASE_URL, **_PG_KWARGS)
    try:
        if truncate:
            # Reap sessions abandoned mid-read by a dropped pooler connection. They
            # sit 'idle in transaction' holding ACCESS SHARE on `sales` and would
            # block our TRUNCATE indefinitely. Only genuinely idle transactions are
            # touched — never a query that is actively running.
            conn.autocommit = True
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT pg_terminate_backend(pid) FROM pg_stat_activity
                    WHERE datname = current_database()
                      AND pid <> pg_backend_pid()
                      AND state = 'idle in transaction'
                      AND state_change < now() - interval '30 seconds'
                      AND query ILIKE '%sales%'
                """)
                reaped = cur.rowcount
                if reaped > 0:
                    print(f"[INFO] Cleared {reaped} stale session(s) holding locks on sales")

        conn.autocommit = False
        with conn.cursor() as cur:
            # Generous ceiling for the COPY; the default (2min) is tight for a bulk load.
            cur.execute("SET statement_timeout = '300s'")
            if truncate:
                # TRUNCATE needs ACCESS EXCLUSIVE, so any in-flight read of `sales`
                # blocks it. Fail fast and say so rather than burning the whole
                # request budget waiting on a lock we are never going to get.
                cur.execute("SET lock_timeout = '20s'")
                try:
                    cur.execute("TRUNCATE TABLE sales RESTART IDENTITY")
                except psycopg2.errors.LockNotAvailable:
                    raise RuntimeError(
                        'Could not lock the sales table — another query is still '
                        'reading it. Wait a minute and sync again.'
                    ) from None
            cur.copy_expert(
                f"COPY sales ({','.join(cols)}) FROM STDIN WITH (FORMAT csv, NULL '')",
                buf,
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print(f"[INFO] Pushed {len(df_sb)} records to Supabase "
          f"({'TRUNCATE + COPY' if truncate else 'COPY append'})")
    return len(df_sb)


def _read_analytics_cache():
    """Fetch precomputed analytics from Supabase. Returns the row dict or None."""
    if not SUPABASE_KEY:
        return None
    try:
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/analytics_cache?select=result,updated_at&id=eq.1",
            headers=_sb_headers(),
            timeout=10,
        )
        if resp.ok:
            rows = resp.json()
            if rows:
                return rows[0]
    except Exception as _e:
        print(f"[INFO] Analytics cache read failed: {_e}")
    return None


def _write_analytics_cache(result):
    """Persist precomputed analytics result to Supabase. Silently ignores errors."""
    if not SUPABASE_KEY:
        return
    try:
        import json as _json
        class _Enc(_json.JSONEncoder):
            def default(self, o):
                if hasattr(o, 'item'): return o.item()
                if hasattr(o, 'tolist'): return o.tolist()
                return str(o)
        safe = _json.loads(_json.dumps(result, cls=_Enc))
        safe.pop('cache_status', None)
        requests.post(
            f"{SUPABASE_URL}/rest/v1/analytics_cache",
            headers={**_sb_headers(), 'Prefer': 'resolution=merge-duplicates'},
            data=_json.dumps({'id': 1, 'result': safe}),
            timeout=30,
        )
        print("[INFO] Analytics cache written to Supabase")
    except Exception as _e:
        print(f"[WARNING] Analytics cache write failed: {_e}")


def _bust_analytics_cache():
    """Delete the analytics cache row so the next request recomputes."""
    if not SUPABASE_KEY:
        return
    try:
        requests.delete(
            f"{SUPABASE_URL}/rest/v1/analytics_cache?id=eq.1",
            headers=_sb_headers(),
            timeout=10,
        )
        print("[INFO] Analytics cache busted")
    except Exception as _e:
        print(f"[WARNING] Analytics cache bust failed: {_e}")


def get_customer_data():
    """Return the full sales dataset from Supabase.

    Supabase is the source of truth. Google Sheets is no longer involved: it was
    only ever a data-entry surface, and keeping the two in step is what caused
    the timeouts, lock contention and date-range mismatches this dashboard used
    to suffer from. New rows arrive via POST /api/upload.

    The on-disk CSV is a last-resort fallback so a transient Supabase blip shows
    slightly stale numbers rather than an error page.
    """
    global cached_data, last_fetch_time

    if cached_data is not None and last_fetch_time is not None:
        if time.time() - last_fetch_time < CACHE_DURATION:
            print("[DEBUG] Returning in-memory cached data")
            return cached_data

    # 2. On-disk working copy.
    #
    # Deliberately preferred over re-reading Supabase. Pulling ~259k rows back
    # out of the pooler is slow and unreliable (measured: erratic, 2k rows can
    # take 30s, and full reads regularly die with 'SSL error: unexpected eof'),
    # so we only pay that cost when we have nothing local. Every write path
    # updates this file alongside Supabase, so it does not drift: uploads append
    # to both, and _refresh_analytics rewrites it.
    if os.path.exists(CACHE_FILE):
        try:
            df = pd.read_csv(CACHE_FILE, low_memory=False)
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            # Rebuild the key rather than trusting the round-tripped column.
            if 'Phone' in df.columns:
                df['Customer_ID'] = _customer_id(df['Phone'])
            if not df.empty:
                cached_data     = df
                last_fetch_time = os.path.getmtime(CACHE_FILE)
                print(f"[INFO] Loaded {len(df)} records from local working copy")
                return df
        except Exception as e:
            print(f"[WARNING] Local working copy unreadable ({e}); reading Supabase")

    if not DATABASE_URL:
        raise RuntimeError(
            "DATABASE_URL is not set — the dashboard cannot reach Supabase. "
            "Set it in your .env locally and in the Vercel environment variables."
        )

    # 3. Refill from Supabase (authoritative, but the slow path).
    try:
        df = _load_from_supabase()
        cached_data     = df
        last_fetch_time = time.time()
        _save_working_copy(df)
        return df
    except Exception as e:
        print(f"[ERROR] Supabase read failed: {e}")
        if cached_data is not None:
            print("[WARNING] Serving stale in-memory data")
            return cached_data
        raise RuntimeError(f"Unable to load data from Supabase: {e}")


def _save_working_copy(df):
    """Persist the working copy that get_customer_data() reads on the next hit."""
    try:
        df.to_csv(CACHE_FILE, index=False)
        print(f"[INFO] Local working copy updated ({len(df)} records)")
    except Exception as e:
        print(f"[WARNING] Could not write local working copy: {e}")


def calculate_overview(df):
    """Calculate overview metrics"""
    
    try:
        # Make a copy to avoid modifying original
        df_work = df.copy()
        
        # Remove rows with NaN totals
        df_work = df_work[df_work['Total'].notna()]

        # Total transactions
        total_transactions = len(df_work)

        # Unique customers
        unique_customers = df_work['Customer_ID'].nunique()

        # Average spend per visit
        df_work['Visit_Date'] = df_work['Date'].dt.date
        visit_spending = df_work.groupby(['Customer_ID', 'Visit_Date'])['Total'].sum().reset_index()
        avg_spend_per_visit = visit_spending['Total'].mean() if len(visit_spending) > 0 else 0
        total_visits = len(visit_spending)

        # Average spend per customer
        customer_spending = df_work.groupby('Customer_ID')['Total'].sum()
        avg_spend_per_customer = customer_spending.mean() if len(customer_spending) > 0 else 0
        
        # Customer purchase frequency - count unique visit days per customer
        customer_visit_days = df_work.groupby('Customer_ID')['Visit_Date'].nunique()
        
        # One-timers and repeat customers (based on different days visited)
        one_timers = (customer_visit_days == 1).sum()
        repeat_customers = (customer_visit_days > 1).sum()
        one_timer_pct = (one_timers / unique_customers * 100) if unique_customers > 0 else 0
        repeat_pct = (repeat_customers / unique_customers * 100) if unique_customers > 0 else 0
        
        # Average lifespan (for repeat customers with visits on different days)
        # Difference between first and last purchase date
        lifespan_df = df_work.groupby('Customer_ID')['Date'].agg(['min', 'max'])
        lifespan_df['lifespan_days'] = (lifespan_df['max'] - lifespan_df['min']).dt.days
        repeat_customers_lifespan = lifespan_df[lifespan_df['lifespan_days'] > 0]
        avg_lifespan = repeat_customers_lifespan['lifespan_days'].mean() if len(repeat_customers_lifespan) > 0 else 0
        
        # Average purchase cycle (for repeat customers)
        # Average time between purchases = Total Lifespan / (Number of Visits - 1)
        # We need visit counts per customer
        # Prepare metrics for cycle calculation
        customer_metrics = lifespan_df.copy()
        customer_metrics['visit_count'] = customer_visit_days
        
            # Optimized Average Purchase Cycle Calculation (Vectorized)
        # 1. Sort by Customer and Date to ensure correct diff calculation
        df_sorted = df_work.sort_values(['Customer_ID', 'Date'])
        
        # 2. Calculate time difference between consecutive rows
        df_sorted['date_diff'] = df_sorted['Date'].diff().dt.days
        
        # 3. Create a mask to identify the start of a new customer block
        # The first record for each customer should not be compared to the previous customer's last record
        mask = df_sorted['Customer_ID'] != df_sorted['Customer_ID'].shift(1)
        
        # 4. Set diff to NaN for the first visit of each customer
        df_sorted.loc[mask, 'date_diff'] = np.nan
        
        # 5. Filter for valid differences (repeat visits)
        valid_diffs = df_sorted.dropna(subset=['date_diff'])
        
        if not valid_diffs.empty:
            # 6. Calculate average cycle per customer first (to match user logic: Mean of (Mean per Customer))
            per_customer_avg = valid_diffs.groupby('Customer_ID')['date_diff'].mean()
            avg_cycle = per_customer_avg.mean()
        else:
            avg_cycle = 0

        total_revenue = df_work['Total'].sum()

        # TWO APPROACHES REPEAT ANALYSIS
        # 1. Transaction-Based: Any customer with > 1 total transaction
        customer_trx_counts = df_work.groupby('Customer_ID').size()
        repeat_customers_trx = (customer_trx_counts > 1).sum()
        repeat_pct_trx = (repeat_customers_trx / unique_customers * 100) if unique_customers > 0 else 0

        # 2. Visit-Day Based: Any customer with > 1 unique visit day (Current logic)
        # Already calculated above as repeat_customers and repeat_pct

        # Date range
        start_date = df_work['Date'].min().strftime('%Y-%m-%d') if not df_work.empty else "N/A"
        end_date = df_work['Date'].max().strftime('%Y-%m-%d') if not df_work.empty else "N/A"
        
        # Calculate Marketing Spend and CAC
        marketing_cost = 0
        cac = 0
        if 'MARKETING EXPENSE' in df_work.columns:
            # Each row carries its own marketing expense value — sum the full column.
            marketing_cost = df_work['MARKETING EXPENSE'].sum()
            cac = marketing_cost / unique_customers if unique_customers > 0 else 0

        return {
            'totalPurchases': int(total_transactions),
            'totalVisits': int(total_visits),
            'uniqueCustomers': int(unique_customers),
            'avgSpendPerVisit': round(float(avg_spend_per_visit), 2) if not np.isnan(avg_spend_per_visit) else 0,
            'avgSpendPerCustomer': round(float(avg_spend_per_customer), 2) if not np.isnan(avg_spend_per_customer) else 0,
            'totalRevenue': round(float(total_revenue), 2),
            'oneTimers': int(one_timers),
            'oneTimerPct': round(float(one_timer_pct), 2),
            'repeatCustomers': int(repeat_customers),
            'repeatPct': round(float(repeat_pct), 2),
            'repeatCustomersTrx': int(repeat_customers_trx),
            'repeatPctTrx': round(float(repeat_pct_trx), 2),
            'avgLifespan': round(float(avg_lifespan), 2) if not np.isnan(avg_lifespan) else 0,
            'avgCycle': float(avg_cycle) if not np.isnan(avg_cycle) else 0,
            'startDate': start_date,
            'endDate': end_date,
            'marketingSpend': float(marketing_cost),
            'cac': float(cac)
        }
    except Exception as e:
        raise Exception(f"Error calculating overview: {str(e)}")

def calculate_visit_interval_distribution(df):
    """Calculate the distribution of days between visits for repeat customers"""
    try:
        df_sorted = df.sort_values(['Customer_ID', 'Date'])
        df_sorted['date_diff'] = df_sorted['Date'].diff().dt.days
        mask = df_sorted['Customer_ID'] != df_sorted['Customer_ID'].shift(1)
        df_sorted.loc[mask, 'date_diff'] = np.nan
        
        valid_diffs = df_sorted.dropna(subset=['date_diff'])['date_diff']
        
        if valid_diffs.empty:
            return []
            
        # Define bins
        bins = [0, 7, 14, 30, 60, 90, float('inf')]
        labels = ['0-7 Days', '8-14 Days', '15-30 Days', '31-60 Days', '61-90 Days', '91+ Days']
        
        dist = pd.cut(valid_diffs, bins=bins, labels=labels, right=True).value_counts().sort_index()
        
        results = []
        for label, count in dist.items():
            results.append({
                'bin': str(label),
                'count': int(count),
                'percentage': round(float(count / len(valid_diffs) * 100), 2)
            })
            
        return results
    except Exception as e:
        print(f"Error calculating interval distribution: {str(e)}")
        return []


def calculate_retention_repeat(period_df, prev_period_df=None, full_df=None):
    """Calculate retention and repeat metrics including spend, lifespan and growth"""
    
    try:
        current_customers_series = period_df['Customer_ID']
        current_customers = {cid for cid in current_customers_series.unique() if pd.notna(cid)}
        current_count = len(current_customers)
        
        # 1. Visit-Day Based Repeat
        period_df_copy = period_df.copy()
        period_df_copy['Visit_Date'] = period_df_copy['Date'].dt.date
        visit_days = period_df_copy.groupby('Customer_ID')['Visit_Date'].nunique()
        repeat_in_period = (visit_days > 1).sum()
        repeat_pct = (repeat_in_period / current_count * 100) if current_count > 0 else 0

        # 2. Transaction-Based Repeat
        trx_counts = period_df.groupby('Customer_ID').size()
        repeat_trx_count = (trx_counts > 1).sum()
        repeat_trx_pct = (repeat_trx_count / current_count * 100) if current_count > 0 else 0
        
        # 3. Avg Spend Per Customer
        total_revenue = period_df['Total'].sum()
        avg_spend = (total_revenue / current_count) if current_count > 0 else 0

        # 4. Avg Lifespan (Days) - Historical lifespan for active customers
        avg_lifespan = 0
        if full_df is not None and not current_customers_series.empty:
            active_ids = current_customers_series.unique()
            # Filter full_df for these specific customers to get their global min/max
            lifespan_data = full_df[full_df['Customer_ID'].isin(active_ids)].groupby('Customer_ID')['Date'].agg(['min', 'max'])
            lifespan_days = (lifespan_data['max'] - lifespan_data['min']).dt.days
            avg_lifespan = lifespan_days.mean() if not lifespan_days.empty else 0

        # 5. % Change (Growth Rate)
        growth_rate = 0
        if prev_period_df is not None and not prev_period_df.empty:
            prev_count = prev_period_df['Customer_ID'].nunique()
            if prev_count > 0:
                growth_rate = ((current_count - prev_count) / prev_count) * 100
        
        # Retention
        if prev_period_df is not None and not prev_period_df.empty:
            prev_customers = {cid for cid in prev_period_df['Customer_ID'].unique() if pd.notna(cid)}
            retained_customers = current_customers.intersection(prev_customers)
            retained_count = len(retained_customers)
            retention_pct = (retained_count / len(prev_customers) * 100) if len(prev_customers) > 0 else 0
        else:
            retained_count = 0
            retention_pct = 0
        
        return {
            'totalCustomers': int(current_count),
            'retainedCustomers': int(retained_count),
            'retentionPct': round(float(retention_pct), 2),
            'repeatCustomers': int(repeat_in_period),
            'repeatPct': round(float(repeat_pct), 2),
            'repeatCustomersTrx': int(repeat_trx_count),
            'repeatPctTrx': round(float(repeat_trx_pct), 2),
            'avgSpendPerCustomer': round(float(avg_spend), 2),
            'avgLifespan': round(float(avg_lifespan), 2),
            'growthRate': round(float(growth_rate), 2)
        }
    except Exception as e:
        raise Exception(f"Error calculating retention: {str(e)}")

def calculate_monthly_repeat_breakdown(df):
    """Calculate breakdown of repeat customers across months"""
    try:
        df_copy = df.dropna(subset=['Date']).copy()
        df_copy['YearMonth'] = df_copy['Date'].dt.to_period('M')
        df_copy['Visit_Date'] = df_copy['Date'].dt.date
        months = sorted(df_copy['YearMonth'].unique())
        
        monthly_breakdown = []
        
        for month in months:
            month_df = df_copy[df_copy['YearMonth'] == month]
            
            # Count unique visit days per customer in this month
            visit_days = month_df.groupby('Customer_ID')['Visit_Date'].nunique()
            
            # Breakdown by customer type
            one_timers_month = (visit_days == 1).sum()
            repeat_customers_month = (visit_days > 1).sum()
            total_customers_month = len(visit_days)
            
            # Calculate percentages
            repeat_pct_month = (repeat_customers_month / total_customers_month * 100) if total_customers_month > 0 else 0
            one_timer_pct_month = (one_timers_month / total_customers_month * 100) if total_customers_month > 0 else 0
            
            # Total revenue for the month
            total_revenue_month = month_df['Total'].sum()

            # Revenue from repeat customers
            repeat_customer_ids = visit_days[visit_days > 1].index
            repeat_revenue = month_df[month_df['Customer_ID'].isin(repeat_customer_ids)]['Total'].sum()
            repeat_revenue_pct = (repeat_revenue / total_revenue_month * 100) if total_revenue_month > 0 else 0
            
            monthly_breakdown.append({
                'period': str(month),
                'totalCustomers': int(total_customers_month),
                'oneTimers': int(one_timers_month),
                'oneTimerPct': round(float(one_timer_pct_month), 2),
                'repeatCustomers': int(repeat_customers_month),
                'repeatPct': round(float(repeat_pct_month), 2),
                'totalRevenue': round(float(total_revenue_month), 2),
                'repeatRevenue': round(float(repeat_revenue), 2),
                'repeatRevenuePct': round(float(repeat_revenue_pct), 2)
            })
        
        return monthly_breakdown
    except Exception as e:
        raise Exception(f"Error calculating monthly repeat breakdown: {str(e)}")

def calculate_semiannual_repeat_breakdown(df):
    """Calculate breakdown of repeat customers across semi-annual periods"""
    try:
        df_copy = df.dropna(subset=['Date']).copy()
        df_copy['Half'] = df_copy['Date'].dt.month.apply(lambda x: 'H1' if x <= 6 else 'H2')
        df_copy['YearHalf'] = df_copy['Date'].dt.year.astype(str) + '-' + df_copy['Half']
        df_copy['Visit_Date'] = df_copy['Date'].dt.date
        
        halves = sorted(df_copy['YearHalf'].unique())
        semi_annual_breakdown = []
        
        for half in halves:
            half_df = df_copy[df_copy['YearHalf'] == half]
            
            # Count unique visit days per customer in this half
            visit_days = half_df.groupby('Customer_ID')['Visit_Date'].nunique()
            
            # Breakdown by customer type
            one_timers_half = (visit_days == 1).sum()
            repeat_customers_half = (visit_days > 1).sum()
            total_customers_half = len(visit_days)
            
            # Calculate percentages
            repeat_pct_half = (repeat_customers_half / total_customers_half * 100) if total_customers_half > 0 else 0
            one_timer_pct_half = (one_timers_half / total_customers_half * 100) if total_customers_half > 0 else 0
            
            # Total revenue for the half
            total_revenue_half = half_df['Total'].sum()

            # Revenue from repeat customers
            repeat_customer_ids = visit_days[visit_days > 1].index
            repeat_revenue = half_df[half_df['Customer_ID'].isin(repeat_customer_ids)]['Total'].sum()
            repeat_revenue_pct = (repeat_revenue / total_revenue_half * 100) if total_revenue_half > 0 else 0
            
            semi_annual_breakdown.append({
                'period': half,
                'totalCustomers': int(total_customers_half),
                'oneTimers': int(one_timers_half),
                'oneTimerPct': round(float(one_timer_pct_half), 2),
                'repeatCustomers': int(repeat_customers_half),
                'repeatPct': round(float(repeat_pct_half), 2),
                'totalRevenue': round(float(total_revenue_half), 2),
                'repeatRevenue': round(float(repeat_revenue), 2),
                'repeatRevenuePct': round(float(repeat_revenue_pct), 2)
            })
        
        return semi_annual_breakdown
    except Exception as e:
        raise Exception(f"Error calculating semi-annual repeat breakdown: {str(e)}")

def calculate_overall_repeat_breakdown(df):
    """Calculate overall breakdown of repeat customers for the entire year"""
    try:
        df_copy = df.copy()
        df_copy['Visit_Date'] = df_copy['Date'].dt.date
        
        # Count unique visit days per customer for the entire year
        visit_days = df_copy.groupby('Customer_ID')['Visit_Date'].nunique()
        
        # Breakdown by customer type
        one_timers_overall = (visit_days == 1).sum()
        repeat_customers_overall = (visit_days > 1).sum()
        total_customers_overall = len(visit_days)
        
        # Calculate percentages
        repeat_pct_overall = (repeat_customers_overall / total_customers_overall * 100) if total_customers_overall > 0 else 0
        one_timer_pct_overall = (one_timers_overall / total_customers_overall * 100) if total_customers_overall > 0 else 0
        
        # Total revenue for the year
        total_revenue_overall = df_copy['Total'].sum()

        # Revenue from repeat customers
        repeat_customer_ids = visit_days[visit_days > 1].index
        repeat_revenue = df_copy[df_copy['Customer_ID'].isin(repeat_customer_ids)]['Total'].sum()
        repeat_revenue_pct = (repeat_revenue / total_revenue_overall * 100) if total_revenue_overall > 0 else 0

        # Additional metrics
        avg_transactions_per_customer = len(df_copy) / total_customers_overall if total_customers_overall > 0 else 0
        avg_spend_per_customer = df_copy.groupby('Customer_ID')['Total'].sum().mean()
        
        # Dynamic period based on data range
        min_year = df_copy['Date'].dt.year.min()
        max_year = df_copy['Date'].dt.year.max()
        period = f"{min_year}" if min_year == max_year else f"{min_year}-{max_year}"
        
        return {
            'period': period,
            'totalCustomers': int(total_customers_overall),
            'oneTimers': int(one_timers_overall),
            'oneTimerPct': round(float(one_timer_pct_overall), 2),
            'repeatCustomers': int(repeat_customers_overall),
            'repeatPct': round(float(repeat_pct_overall), 2),
            'totalRevenue': round(float(total_revenue_overall), 2),
            'repeatRevenue': round(float(repeat_revenue), 2),
            'repeatRevenuePct': round(float(repeat_revenue_pct), 2),
            'avgTransactionsPerCustomer': round(float(avg_transactions_per_customer), 2),
            'avgSpendPerCustomer': round(float(avg_spend_per_customer), 2)
        }
    except Exception as e:
        raise Exception(f"Error calculating overall repeat breakdown: {str(e)}")

def _calculate_trend_data(df_copy, period_column):
    """Generic function to calculate period-over-period trend data (Monthly, Quarterly, Semi-Annual, etc)"""
    try:
        results = []
        all_seen_so_far = set()
        prev_period_df = None
        prev_result = None
        
        # Pre-calculate min/max dates for vectorized lifespan
        customer_min_date = pd.Series(dtype='datetime64[ns]')
        customer_max_date = pd.Series(dtype='datetime64[ns]')
        
        # Optimize: GroupBy is much faster than manual filtering in a loop
        for period, period_df in df_copy.groupby(period_column, sort=True):
            # Get unique customers and ensure no NaN/Null values crash calculations
            if period_df.empty: continue
            # Get unique customers and ensure no NaN/Null values crash calculations
            current_customers = {cid for cid in period_df['Customer_ID'].unique() if pd.notna(cid)}
            current_count = len(current_customers)
            
            if current_count == 0:
                continue

            # 1. Retention (from previous period)
            retained_count = 0
            retention_pct = 0
            if prev_period_df is not None and not prev_period_df.empty:
                prev_customers = set(prev_period_df['Customer_ID'].unique())
                retained_customers = current_customers.intersection(prev_customers)
                retained_count = len(retained_customers)
                retention_pct = (retained_count / len(prev_customers) * 100) if len(prev_customers) > 0 else 0
                
            # 2. New Customers (First time seen in this dataset)
            new_customers = current_customers - all_seen_so_far
            new_count = len(new_customers)
            new_pct = (new_count / current_count * 100) if current_count > 0 else 0
            
            # 3. Repeat (Visit-day based)
            period_df_copy = period_df.copy()
            period_df_copy['Visit_Date'] = period_df_copy['Date'].dt.date
            visit_days = period_df_copy.groupby('Customer_ID')['Visit_Date'].nunique()
            repeat_count = (visit_days > 1).sum()
            repeat_pct = (repeat_count / current_count * 100) if current_count > 0 else 0
            
            # 4. Repeat Transaction Based
            trx_counts = period_df.groupby('Customer_ID').size()
            repeat_trx_count = (trx_counts > 1).sum()
            repeat_trx_pct = (repeat_trx_count / current_count * 100) if current_count > 0 else 0
            
            # 5. Average Spend
            total_revenue = period_df['Total'].sum()
            avg_spend = total_revenue / current_count if current_count > 0 else 0
            
            # 6. Average Lifespan (Cumulative Tracking)
            # Efficiently update min/max dates for all customers in this period
            item_stats = period_df.groupby('Customer_ID')['Date'].agg(['min', 'max'])
            
            # Combine current stats with historical - much faster than iteration/dicts
            customer_min_date = pd.concat([customer_min_date, item_stats['min']]).groupby(level=0).min()
            customer_max_date = pd.concat([customer_max_date, item_stats['max']]).groupby(level=0).max()
                
            # Vectorized lifespan calculation
            lifespan_days = (customer_max_date - customer_min_date).dt.days
            repeat_lifespans = lifespan_days[lifespan_days > 0]
            avg_lifespan = float(repeat_lifespans.mean()) if not repeat_lifespans.empty else 0
            
            # Extract Revenue for backward compatibility with semi-annual logic
            repeat_customer_ids = visit_days[visit_days > 1].index
            repeat_revenue = period_df[period_df['Customer_ID'].isin(repeat_customer_ids)]['Total'].sum()
            repeat_revenue_pct = (repeat_revenue / total_revenue * 100) if total_revenue > 0 else 0
                
            result = {
                'period': str(period),
                'totalCustomers': int(current_count),
                'newCustomers': int(new_count),
                'newPct': round(float(new_pct), 2),
                'retainedCustomers': int(retained_count),
                'retentionPct': round(float(retention_pct), 2),
                'repeatCustomers': int(repeat_count),
                'repeatPct': round(float(repeat_pct), 2),
                'repeatCustomersTrx': int(repeat_trx_count),
                'repeatPctTrx': round(float(repeat_trx_pct), 2),
                'avgSpendPerCustomer': round(float(avg_spend), 2),
                'avgLifespan': round(float(avg_lifespan), 2),
                'growthRate': 0,  # Default for first period
                'marketingSpend': 0,
                'cac': 0,
                'cacOverall': 0,
                'totalRevenue': round(float(total_revenue), 2),
                'repeatRevenue': round(float(repeat_revenue), 2),
                'repeatRevenuePct': round(float(repeat_revenue_pct), 2)
            }
            
            # 7. CAC Calculation from Sheet Column
            if 'MARKETING EXPENSE' in period_df.columns:
                # Each row carries its own marketing expense value — sum the full column.
                marketing_cost = period_df['MARKETING EXPENSE'].sum()
                result['marketingSpend'] = float(marketing_cost)
                result['cac'] = round(marketing_cost / new_count, 2) if new_count > 0 else 0
                result['cacOverall'] = round(marketing_cost / current_count, 2) if current_count > 0 else 0
            
            # Period-over-Period Growth & Comparative Values
            def get_growth(curr, prev):
                if prev is None or prev == 0: return 0
                return round(((curr - prev) / prev * 100), 2)

            if prev_result:
                result['growthRate'] = get_growth(result['totalCustomers'], prev_result['totalCustomers'])
                result['prev_repeatCustomers'] = prev_result['repeatCustomers']
                result['repeatGrowth'] = get_growth(result['repeatCustomers'], prev_result['repeatCustomers'])
                
                result['prev_avgSpend'] = prev_result['avgSpendPerCustomer']
                result['avgSpendGrowth'] = get_growth(result['avgSpendPerCustomer'], prev_result['avgSpendPerCustomer'])
                
                result['prev_avgLifespan'] = prev_result['avgLifespan']
                result['avgLifespanGrowth'] = get_growth(result['avgLifespan'], prev_result['avgLifespan'])
            else:
                result['prev_repeatCustomers'] = 0
                result['repeatGrowth'] = 0
                result['prev_avgSpend'] = 0
                result['avgSpendGrowth'] = 0
                result['prev_avgLifespan'] = 0
                result['avgLifespanGrowth'] = 0
            
            results.append(result)
            all_seen_so_far.update(current_customers)
            prev_period_df = period_df
            prev_result = result
            
        return results
    except Exception as e:
        raise Exception(f"Error calculating trend data: {str(e)}")

def _prepare_working_df(df):
    """Sort and add all necessary temporal columns once to avoid redundant expensive calculations."""
    df_copy = df.copy().sort_values('Date')
    df_copy['Visit_Date'] = df_copy['Date'].dt.date
    df_copy['YearMonth'] = df_copy['Date'].dt.to_period('M')
    df_copy['YearQuarter'] = df_copy['Date'].dt.to_period('Q')
    df_copy['YearHalf'] = df_copy['Date'].dt.year.astype(str) + '-H' + (df_copy['Date'].dt.month.le(6).map({True: '1', False: '2'}))
    df_copy['Year'] = df_copy['Date'].dt.year
    # ISO Week format: YYYY-W##
    df_copy['YearWeek'] = df_copy['Date'].dt.isocalendar().year.astype(str) + '-W' + df_copy['Date'].dt.isocalendar().week.astype(str).str.zfill(2)
    return df_copy

def calculate_monthly_data(df):
    """Calculate month-to-month metrics using pre-prepared df"""
    try:
        # Assumes df already has YearMonth
        return _calculate_trend_data(df, 'YearMonth')
    except Exception as e:
        raise Exception(f"Error calculating monthly data: {str(e)}")


def calculate_weekly_data(df):
    """Calculate weekly metrics from week 1 of 2026 to current week"""
    try:
        # Filter for data from 2026 onwards
        df_filtered = df[df['Date'].dt.year >= 2026].copy()
        
        if df_filtered.empty:
            return []
        
        # Assumes df already has YearWeek
        weekly_data = _calculate_trend_data(df_filtered, 'YearWeek')
        return weekly_data
    except Exception as e:
        raise Exception(f"Error calculating weekly data: {str(e)}")



def calculate_cumulative_retention(df, start_date='2025-04-01'):
    """
    Calculate retention and repeat growth from a fixed start date.
    Tracks 'returning customers' and 'total repeaters' found so far.
    """
    try:
        df_copy = df.copy()
        df_copy['Date'] = pd.to_datetime(df_copy['Date'])
        df_copy['Visit_Date'] = df_copy['Date'].dt.date
        start_date_dt = pd.to_datetime(start_date)
        
        df_filtered = df_copy[df_copy['Date'] >= start_date_dt].copy()
        if df_filtered.empty: return []
            
        df_filtered['YearMonth'] = df_filtered['Date'].dt.to_period('M')
        months = sorted(df_filtered['YearMonth'].unique())
        
        cumulative_results = []
        seen_customer_visit_days = {} # Customer_ID -> set of Visit_Dates
        
        for i, month in enumerate(months):
            month_df = df_filtered[df_filtered['YearMonth'] == month]
            total_revenue = month_df['Total'].sum()

            # Customers in this month
            current_month_visits = month_df.groupby('Customer_ID')['Visit_Date'].unique().to_dict()
            current_customers = set(current_month_visits.keys())
            
            # Retained: Seen in ANY previous month
            all_seen_so_far = set(seen_customer_visit_days.keys())
            retained_customers = current_customers.intersection(all_seen_so_far)
            retained_count = len(retained_customers)
            retention_pct = (retained_count / len(current_customers) * 100) if len(current_customers) > 0 else 0
            
            # New: Not seen before
            new_customers = current_customers - all_seen_so_far
            new_count = len(new_customers)
            new_pct = (new_count / len(current_customers) * 100) if len(current_customers) > 0 else 0

            # Update master tracker
            for cid, vdays in current_month_visits.items():
                if cid not in seen_customer_visit_days:
                    seen_customer_visit_days[cid] = set()
                seen_customer_visit_days[cid].update(vdays)
            
            # Cumulative Stats
            total_unique_so_far = len(seen_customer_visit_days)
            repeaters_so_far = sum(1 for cid, vdays in seen_customer_visit_days.items() if len(vdays) > 1)
            cumulative_repeat_pct = (repeaters_so_far / total_unique_so_far * 100) if total_unique_so_far > 0 else 0
            
            retained_revenue = month_df[month_df['Customer_ID'].isin(retained_customers)]['Total'].sum()
            retained_revenue_pct = (retained_revenue / total_revenue * 100) if total_revenue > 0 else 0
            
            cumulative_results.append({
                'period': str(month),
                'totalCustomers': int(len(current_customers)),
                'retainedCustomers': int(retained_count),
                'retentionPct': round(float(retention_pct), 2),
                'newCustomers': int(new_count),
                'newPct': round(float(new_pct), 2),
                'totalRevenue': round(float(total_revenue), 2),
                'retainedRevenue': round(float(retained_revenue), 2),
                'retainedRevenuePct': round(float(retained_revenue_pct), 2),
                'cumulativeRepeatCount': int(repeaters_so_far),
                'cumulativeRepeatPct': round(float(cumulative_repeat_pct), 2),
                'cumulativeUniqueTotal': int(total_unique_so_far)
            })
            
        return cumulative_results
    except Exception as e:
        print(f"[ERROR] Error in calculate_cumulative_retention: {str(e)}")
        return []

def calculate_quarterly_data(df):
    """Calculate quarterly metrics using pre-prepared df"""
    try:
        # Assumes df already has YearQuarter
        return _calculate_trend_data(df, 'YearQuarter')
    except Exception as e:
        raise Exception(f"Error calculating quarterly data: {str(e)}")

def calculate_overall_performance(df):
    """Calculate overall performance metrics for the entire data period"""
    try:
        # Assumes df already has Visit_Date
        df_copy = df
        
        # Total metrics
        total_customers = df_copy['Customer_ID'].nunique()
        total_transactions = len(df_copy)
        total_revenue = df_copy['Total'].sum()

        # Visit metrics
        total_visits = len(df_copy.groupby(['Customer_ID', 'Visit_Date']))
        avg_spend_per_visit = df_copy.groupby(['Customer_ID', 'Visit_Date'])['Total'].sum().mean()

        # Customer types
        visit_days = df_copy.groupby('Customer_ID')['Visit_Date'].nunique()
        one_timers = (visit_days == 1).sum()
        repeat_customers = (visit_days > 1).sum()
        one_timer_pct = (one_timers / total_customers * 100) if total_customers > 0 else 0
        repeat_pct = (repeat_customers / total_customers * 100) if total_customers > 0 else 0

        # Revenue breakdown
        repeat_customer_ids = visit_days[visit_days > 1].index
        repeat_revenue = df_copy[df_copy['Customer_ID'].isin(repeat_customer_ids)]['Total'].sum()
        repeat_revenue_pct = (repeat_revenue / total_revenue * 100) if total_revenue > 0 else 0

        # Average customer metrics
        avg_spend_per_customer = df_copy.groupby('Customer_ID')['Total'].sum().mean()
        # Average lifespan (for repeat customers with visits on different days)
        # Difference between first and last purchase date
        lifespan_df = df_copy.groupby('Customer_ID')['Date'].agg(['min', 'max'])
        lifespan_df['lifespan_days'] = (lifespan_df['max'] - lifespan_df['min']).dt.days
        repeat_customers_lifespan = lifespan_df[lifespan_df['lifespan_days'] > 0]
        avg_lifespan = repeat_customers_lifespan['lifespan_days'].mean() if len(repeat_customers_lifespan) > 0 else 0
        
        # Performance status
        repeat_status = '✓' if 20 <= repeat_pct <= 30 else '✗'
        
        # Calculate Marketing Spend and CAC
        marketing_cost = 0
        cac = 0
        if 'MARKETING EXPENSE' in df_copy.columns:
            marketing_cost = df_copy.groupby(df_copy['Date'].dt.date)['MARKETING EXPENSE'].max().sum()
            cac = marketing_cost / total_customers if total_customers > 0 else 0

        return {
            'totalCustomers': int(total_customers),
            'totalTransactions': int(total_transactions),
            'totalVisits': int(total_visits),
            'totalRevenue': round(float(total_revenue), 2),
            'avgSpendPerVisit': round(float(avg_spend_per_visit), 2) if not np.isnan(avg_spend_per_visit) else 0,
            'avgSpendPerCustomer': round(float(avg_spend_per_customer), 2) if not np.isnan(avg_spend_per_customer) else 0,
            'oneTimers': int(one_timers),
            'oneTimerPct': round(float(one_timer_pct), 2),
            'repeatCustomers': int(repeat_customers),
            'repeatPct': round(float(repeat_pct), 2),
            'repeatRevenue': round(float(repeat_revenue), 2),
            'repeatRevenuePct': round(float(repeat_revenue_pct), 2),
            'avgLifespan': round(float(avg_lifespan), 2) if not np.isnan(avg_lifespan) else 0,
            'performanceStatus': repeat_status,
            'marketingSpend': float(marketing_cost),
            'cac': float(cac)
        }
    except Exception as e:
        raise Exception(f"Error calculating overall performance: {str(e)}")

def calculate_semiannual_performance(df):
    """Calculate semi-annual metrics with revenue breakdown using pre-prepared df"""
    try:
        # Assumes df already has YearHalf
        return _calculate_trend_data(df, 'YearHalf')
    except Exception as e:
        raise Exception(f"Error calculating semi-annual data: {str(e)}")

def calculate_yearly_data(df):
    """Calculate yearly metrics using pre-prepared df"""
    try:
        # Assumes df already has Year
        return _calculate_trend_data(df, 'Year')
    except Exception as e:
        raise Exception(f"Error calculating yearly data: {str(e)}")

def calculate_regional_data(df):
    """Calculate regional performance metrics"""
    try:
        if df is None or df.empty:
            return []
            
        regions = {}
        
        # Group shops by region - only include shops that exist in the data
        available_shops = df['Shop'].unique() if 'Shop' in df.columns else []
        
        for shop_name in available_shops:
            if shop_name in SHOP_REGION_MAP:
                region = SHOP_REGION_MAP[shop_name]
                if region not in regions:
                    regions[region] = []
                regions[region].append(shop_name)
        
        regional_results = []
        
        for region, shops_in_region in regions.items():
            # Filter data for shops in this region
            region_df = df[df['Shop'].isin(shops_in_region)]
            
            if region_df.empty:
                continue
            
            # Calculate aggregated metrics for the region
            region_metrics = calculate_overall_performance(region_df)
            region_overview = calculate_overview(region_df)
            
            # Add region-specific data
            regional_results.append({
                'region': region,
                'color': REGION_COLORS.get(region, '#95a5a6'),
                'shops': shops_in_region,
                'totalShops': len(shops_in_region),
                'totalCustomers': region_metrics['totalCustomers'],
                'repeatCustomers': region_metrics['repeatCustomers'],
                'repeatPct': region_metrics['repeatPct'],
                'totalRevenue': region_metrics['totalRevenue'],
                'repeatRevenuePct': region_metrics['repeatRevenuePct'],
                'avgSpendPerCustomer': region_metrics['avgSpendPerCustomer'],
                'performanceStatus': region_metrics['performanceStatus'],
                'oneTimerPct': region_overview['oneTimerPct'],
                'avgLifespan': region_overview['avgLifespan']
            })
        
        # Sort by total revenue descending
        regional_results.sort(key=lambda x: x['totalRevenue'], reverse=True)
        
        return regional_results
    except Exception as e:
        print(f"[ERROR] Error in calculate_regional_data: {str(e)}")
        return []

def calculate_gender_performance(df):
    """Calculate performance metrics by gender"""
    try:
        if 'Gender' not in df.columns:
            return []
        
        gender_results = []
        genders = df['Gender'].unique()
        
        for gender in genders:
            if pd.isna(gender) or gender == '':
                continue
                
            gender_df = df[df['Gender'] == gender]
            
            if gender_df.empty:
                continue
            
            # Calculate metrics
            total_customers = gender_df['Customer_ID'].nunique()
            total_revenue = gender_df['Total'].sum()
            avg_spend = total_revenue / total_customers if total_customers > 0 else 0
            total_transactions = len(gender_df)
            avg_transactions_per_customer = total_transactions / total_customers if total_customers > 0 else 0
            
            # Repeat customers
            visit_days = gender_df.groupby('Customer_ID')['Date'].nunique()
            repeat_customers = (visit_days > 1).sum()
            repeat_pct = (repeat_customers / total_customers * 100) if total_customers > 0 else 0
            
            gender_results.append({
                'gender': gender,
                'totalCustomers': int(total_customers),
                'totalRevenue': round(float(total_revenue), 2),
                'avgSpendPerCustomer': round(float(avg_spend), 2),
                'avgTransactionsPerCustomer': round(float(avg_transactions_per_customer), 2),
                'repeatCustomers': int(repeat_customers),
                'repeatPct': round(float(repeat_pct), 2),
                'totalTransactions': int(total_transactions)
            })
        
        return gender_results
    except Exception as e:
        print(f"[ERROR] Error in calculate_gender_performance: {str(e)}")
        return []

def calculate_product_performance(df):
    """Calculate top products by revenue"""
    try:
        # Check for product column - could be 'Product', 'Item', 'Product Name', etc.
        product_column = None
        possible_columns = ['Product', 'Item', 'Product Name', 'Item Name', 'ProductName', 'ItemName', 'Items', 'Description', 'Specifics']
        
        for col in possible_columns:
            if col in df.columns:
                product_column = col
                break
        
        if product_column is None:
            print("[WARNING] No product column found")
            return []
        
        # Group by product and calculate metrics
        product_stats = df.groupby(product_column).agg(
            totalRevenue=('Total', 'sum'),
            totalSales=('Price', 'count'),
            avgPrice=('Price', 'mean'),
            uniqueCustomers=('Customer_ID', 'nunique')
        ).reset_index()
        
        # Sort by total revenue and get top 10
        top_products = product_stats.nlargest(10, 'totalRevenue')
        
        original_prices = {
            'JUMBO': 3600, 'MAN BAG': 1600, 'ANTITHEFT': 2600, 'CODE 3': 2100,
            'Standard Travel': 2600, 'SAFIRI TRAVEL': 3600, 'FABELA': 2600,
            'KAI': 2800, 'ELYSE': 3100, 'LOLA': 2100, 'MEGA': 3000, 'SCHOOL BAG': 2100
        }

        product_results = []
        for _, row in top_products.iterrows():
            product_name = row[product_column]
            # Fuzzy match or direct match for original price
            orig_price = original_prices.get(product_name, 0)
            avg_sold_price = row['avgPrice']
            
            # If explicit match not found, try case-insensitive or partial
            if orig_price == 0:
                for k, v in original_prices.items():
                    if k.lower() in str(product_name).lower():
                        orig_price = v
                        break
            
            discount_impact = 0
            if orig_price > 0:
                discount_impact = (orig_price - avg_sold_price) * row['totalSales']

            product_results.append({
                'product': product_name,
                'totalRevenue': round(float(row['totalRevenue']), 2),
                'totalSales': int(row['totalSales']),
                'avgPrice': round(float(row['avgPrice']), 2),
                'uniqueCustomers': int(row['uniqueCustomers']),
                'originalPrice': orig_price,
                'discountImpact': round(float(discount_impact), 2)
            })
        
        return product_results
    except Exception as e:
        print(f"[ERROR] Error in calculate_product_performance: {str(e)}")
        return []

def analyze_combos_and_affinity(df):
    """Analyze pre-defined combos (+) and market basket affinity"""
    try:
        # Check for product column
        product_column = None
        possible_columns = ['Product', 'Item', 'Product Name', 'Item Name', 'ProductName', 'ItemName', 'Items', 'Description', 'Specifics']
        for col in possible_columns:
            if col in df.columns:
                product_column = col
                break
        
        if not product_column: return {'combos': [], 'affinity': []}

        # 1. Analyze Existing Combos (with '+', 'Buy', 'Combo', 'Bundle')
        combo_keywords = r'\+|buy\s+.*get\s+.*|combo|bundle|set'
        combo_mask = df[product_column].astype(str).str.contains(combo_keywords, case=False, regex=True, na=False)
        combos_df = df[combo_mask]
        
        if not combos_df.empty:
            combo_stats = combos_df.groupby(product_column).agg(
                Revenue=('Total', 'sum'),
                SalesCount=('Price', 'count'),
                UniqueCustomers=('Customer_ID', 'nunique')
            ).reset_index()
            combo_stats.rename(columns={product_column: 'ComboName'}, inplace=True)
            top_combos = combo_stats.sort_values('Revenue', ascending=False).head(10).to_dict('records')
        else:
            top_combos = []

        # 2. Market Basket Analysis (Affinity)
        # Group by transaction (Customer + Date)
        df_copy = df.copy()
        df_copy['Visit_Date'] = df_copy['Date'].dt.date
        transactions = df_copy.groupby(['Customer_ID', 'Visit_Date'])[product_column].unique()
        
        # Filter for transactions with > 1 item
        multi_item_txns = transactions[transactions.apply(len) > 1]
        
        from collections import Counter
        pair_counts = Counter()
        triplet_counts = Counter()
        import itertools

        # Limit to the most recent 15,000 multi-item transactions for performance
        # on large datasets, while still being representative.
        recent_txns = multi_item_txns.tail(15000)

        for items in recent_txns:
            # Sort items to ensure (A, B) is same as (B, A)
            sorted_items = sorted([str(i) for i in items])
            
            # Generate pairs (Only for reasonably sized baskets to avoid exponential explosion)
            if len(sorted_items) <= 10:
                pair_counts.update(itertools.combinations(sorted_items, 2))
                
                # Generate triplets
                if len(sorted_items) <= 6:
                    triplet_counts.update(itertools.combinations(sorted_items, 3))
        
        # Convert to list and sort
        affinity_results = []
        
        # Top Pairs
        for pair, count in sorted(pair_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
            affinity_results.append({
                'items': ' + '.join(pair),
                'type': 'Pair',
                'frequency': count
            })
            
        # Top Triplets
        for triplet, count in sorted(triplet_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
            affinity_results.append({
                'items': ' + '.join(triplet),
                'type': 'Triplet',
                'frequency': count
            })

        return {
            'topCombos': top_combos,
            'affinity': affinity_results
        }

    except Exception as e:
        print(f"[ERROR] Affinity Analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'combos': [], 'affinity': []}

def calculate_regional_top_products(df):
    """Get top 5 products per region"""
    try:
        # Product Col Check
        product_column = None
        for col in ['Product', 'Item', 'Product Name', 'Item Name', 'Items', 'Description']:
            if col in df.columns: product_column = col; break
        if not product_column: return {}

        regional_products = {}
        
        df_work = df.copy()
        # Ensure Region exists
        if 'Region' not in df_work.columns and 'Shop' in df_work.columns:
             df_work['Region'] = df_work['Shop'].map(SHOP_REGION_MAP)
        
        if 'Region' not in df_work.columns: return {}

        for region in df_work['Region'].unique():
            if not region or pd.isna(region): continue
            
            region_df = df_work[df_work['Region'] == region]
            
            # Group by Product Revenue
            stats = region_df.groupby(product_column)['Total'].sum().reset_index()
            top_5 = stats.sort_values('Total', ascending=False).head(5)

            products = []
            for _, row in top_5.iterrows():
                products.append({
                    'name': row[product_column],
                    'revenue': float(row['Total'])
                })
            
            regional_products[region] = products
            
        return regional_products

    except Exception as e:
        print(f"[ERROR] Regional Products: {str(e)}")
        return {}

def calculate_top_shops_by_region(df):
    """Calculate top 5 shops by revenue for each region"""
    try:
        if 'Shop' not in df.columns:
            return {}
        
        region_top_shops = {}
        
        # Get available shops and their regions
        available_shops = df['Shop'].unique()
        
        for shop in available_shops:
            if shop in SHOP_REGION_MAP:
                region = SHOP_REGION_MAP[shop]
                if region not in region_top_shops:
                    region_top_shops[region] = []
                
                shop_df = df[df['Shop'] == shop]
                total_revenue = shop_df['Total'].sum()
                total_customers = shop_df['Customer_ID'].nunique()
                total_transactions = len(shop_df)
                
                region_top_shops[region].append({
                    'shop': shop,
                    'totalRevenue': round(float(total_revenue), 2),
                    'totalCustomers': int(total_customers),
                    'totalTransactions': int(total_transactions),
                    'avgSpendPerCustomer': round(float(total_revenue / total_customers), 2) if total_customers > 0 else 0
                })
        
        # Sort and get top 6 for each region
        for region in region_top_shops:
            region_top_shops[region].sort(key=lambda x: x['totalRevenue'], reverse=True)
            region_top_shops[region] = region_top_shops[region][:6]
        
        return region_top_shops
    except Exception as e:
        print(f"[ERROR] Error in calculate_top_shops_by_region: {str(e)}")
        return {}

def calculate_monthly_loyalty_trends(df, target_shop):
    """Calculate month-over-month trends for Target-Only vs Cross-Shop customers"""
    try:
        if df.empty or 'Shop' not in df.columns:
            return []
            
        df_copy = df.dropna(subset=['Date']).copy()
        df_copy['YearMonth'] = df_copy['Date'].dt.to_period('M')
        months = sorted(df_copy['YearMonth'].unique())
        
        # Pre-identify all shops visited by each customer globally in this dataset
        # This determines if they are "Cross-Shop" or "Target-Only" once and for all
        cust_shops = df.groupby('Customer_ID')['Shop'].unique()
        is_target_only = cust_shops.apply(lambda x: len(x) == 1 and x[0] == target_shop)
        target_only_ids = set(is_target_only[is_target_only].index)
        
        trends = []
        for month in months:
            month_df = df_copy[df_copy['YearMonth'] == month]
            target_ids = set(month_df[month_df['Shop'] == target_shop]['Customer_ID'].unique())
            
            if not target_ids:
                continue
                
            total = len(target_ids)
            target_only = len(target_ids.intersection(target_only_ids))
            cross_shop = total - target_only
            
            trends.append({
                'month': str(month),
                'total': total,
                'targetOnly': target_only,
                'targetOnlyPct': round((target_only / total * 100), 1) if total > 0 else 0,
                'crossShop': cross_shop,
                'crossShopPct': round((cross_shop / total * 100), 1) if total > 0 else 0
            })
            
        return trends
    except Exception as e:
        print(f"[ERROR] In loyalty trends for {target_shop}: {e}")
        return []

def calculate_shop_loyalty_analysis(df, target_shop, logic='cross-shop', full_df=None):
    """Classifies target shop customers as new vs existing based on selected logic.
       logic='cross-shop': New(Shop-Only) vs Existing(from other stores)
       logic='internal': New(First visit to shop) vs Existing(Historical visitor)
    """
    try:
        # Use full_df for global context if provided, else use df
        context_df = full_df if full_df is not None else df
        
        if 'Shop' not in context_df.columns:
            return {'error': 'Shop column not found', 'totalCustomers': 0}
            
        # 1. Identify all customers who visited the target shop in the CURRENT (filtered) df
        # Case-insensitive lookup to be safe
        target_customer_ids = df[df['Shop'].str.lower() == target_shop.lower()]['Customer_ID'].unique()
        actual_shop_name = df[df['Shop'].str.lower() == target_shop.lower()]['Shop'].iloc[0] if len(target_customer_ids) > 0 else target_shop
        
        if len(target_customer_ids) == 0:
            return {
                'targetShop': target_shop,
                'totalCustomers': 0, 'newCustomers': 0, 'existingCustomers': 0,
                'sourceShops': [], 'sourceRegions': []
            }
        
        # 2. Filter main DF to ONLY these customers for faster processing
        customer_df = df[df['Customer_ID'].isin(target_customer_ids)].copy()
        
        # 3. Identify New vs Existing based on logic
        # We check only the relevant customers
        if logic == 'internal':
            # NEW if they have only 1 transaction on record (this period) AT THIS SHOP
            shop_visits = customer_df[customer_df['Shop'].str.lower() == target_shop.lower()].groupby('Customer_ID', sort=False).size()
            new_ids = shop_visits[shop_visits == 1].index
            existing_ids = shop_visits[shop_visits > 1].index
        else:
            # logic == 'cross-shop'
            # NEW if only 1 shop total in their history (within the current period)
            shop_counts = customer_df.groupby('Customer_ID', sort=False)['Shop'].nunique()
            new_ids = shop_counts[shop_counts == 1].index
            existing_ids = shop_counts[shop_counts > 1].index
        
        target_customer_ids = customer_df[customer_df['Shop'].str.lower() == target_shop.lower()]['Customer_ID'].unique()
        # Ensure we filter IDs specifically to those visiting target_shop
        new_ids = [cid for cid in new_ids if cid in target_customer_ids]
        existing_ids = [cid for cid in existing_ids if cid in target_customer_ids]
        
        # 4. Details for New Customers (Target-Only)
        new_df = customer_df[customer_df['Customer_ID'].isin(new_ids)]
        new_stats = new_df.groupby('Customer_ID').agg(
            totalPurchases=('Price', 'count'),
            totalRevenue=('Total', 'sum'),
            firstPurchaseDate=('Date', 'min')
        ).reset_index()
        
        # 5. Details for Existing Customers (Cross-shoppers)
        existing_df = customer_df[customer_df['Customer_ID'].isin(existing_ids)]
        
        # Find other shops visited by these customers
        other_visits = existing_df[existing_df['Shop'] != target_shop][['Customer_ID', 'Shop']].drop_duplicates()
        source_shop_counts = other_visits['Shop'].value_counts()
        
        # Region counts
        other_visits['Region'] = other_visits['Shop'].map(lambda x: SHOP_REGION_MAP.get(x, 'Unknown'))
        source_region_counts = other_visits['Region'].value_counts()
        
        # Revenue in target shop for cross-shoppers
        target_rev_existing = existing_df[existing_df['Shop'] == target_shop].groupby('Customer_ID')['Total'].sum()
        
        # First shop visited (absolute first in database)
        first_visits = existing_df.sort_values('Date').groupby('Customer_ID').first()[['Shop']]
        
        # Calculate Revenue and Repeat metrics specifically for the target shop
        target_sales = customer_df[customer_df['Shop'].str.lower() == target_shop.lower()]
        rev_new = target_sales[target_sales['Customer_ID'].isin(new_ids)]['Total'].sum()
        rev_existing = target_sales[target_sales['Customer_ID'].isin(existing_ids)]['Total'].sum()
        total_rev = target_sales['Total'].sum()

        # Notice we are calculating repeat rates based strictly on visits to THIS target shop
        target_sales_dates = target_sales.copy()
        target_sales_dates['_visit_date'] = target_sales_dates['Date'].dt.date
        target_visit_counts = target_sales_dates.groupby('Customer_ID')['_visit_date'].nunique()
        
        new_repeat_count = sum(target_visit_counts[target_visit_counts.index.isin(new_ids)] > 1)
        existing_repeat_count = sum(target_visit_counts[target_visit_counts.index.isin(existing_ids)] > 1)
        overall_repeat_pct = round((new_repeat_count + existing_repeat_count) / max(1, len(target_customer_ids)) * 100, 1)
        
        other_shops_per_cust = other_visits.groupby('Customer_ID')['Shop'].apply(list).to_dict()

        # 6. Format Results
        source_shops = [{'shop': s, 'customerCount': int(c), 
                        'percentage': round((c/max(1, len(existing_ids))*100), 2)} 
                        for s, c in source_shop_counts.items()]
        
        source_regions = [{'region': r, 'customerCount': int(c), 
                          'percentage': round((c/max(1, len(existing_ids))*100), 2)} 
                          for r, c in source_region_counts.items()]
        
        # 7. Internal loyalty (Repeat visitors strictly to THIS shop)
        # Identify customers with >= 2 unique visit dates strictly at this shop
        # Note: .dt must be applied BEFORE groupby, not after
        _target_shop_df = customer_df[customer_df['Shop'] == target_shop].copy()
        _target_shop_df['_visit_date'] = _target_shop_df['Date'].dt.date
        target_visits = _target_shop_df.groupby('Customer_ID')['_visit_date'].nunique()
        internal_repeat_ids = target_visits[target_visits >= 2].index
        source_regions = [
            {'region': r, 'customerCount': int(c), 'percentage': round(c / max(1, len(existing_ids)) * 100, 1)}
            for r, c in source_region_counts.items()
        ]

        # 7. Internal loyalty detail list (top 50 by spend)
        internal_repeat_stats = _target_shop_df[
            _target_shop_df['Customer_ID'].isin(internal_repeat_ids)
        ].groupby('Customer_ID').agg(
            visits=('Date', 'nunique'),
            totalSpend=('Total', 'sum'),
            lastVisit=('Date', 'max')
        ).sort_values('totalSpend', ascending=False).head(50).reset_index()

        internal_loyalty_details = [
            {
                'customerId': row['Customer_ID'],
                'visits': int(row['visits']),
                'totalSpend': round(float(row['totalSpend']), 2),
                'lastVisit': row['lastVisit'].strftime('%Y-%m-%d')
            }
            for _, row in internal_repeat_stats.iterrows()
        ]

        # 8. Cross-shop loyalty detail list (top 50 by spend)
        cross_shop_stats = existing_df.groupby('Customer_ID').agg(
            totalSpendAtTarget=('Total', 'sum'),
            lastVisitAtTarget=('Date', 'max')
        ).sort_values('totalSpendAtTarget', ascending=False).head(50)

        cross_loyalty_details = [
            {
                'customerId': cid,
                'otherShops': other_shops_per_cust.get(cid, []),
                'totalSpendAtTarget': round(float(row['totalSpendAtTarget']), 2),
                'lastVisitAtTarget': row['lastVisitAtTarget'].strftime('%Y-%m-%d')
            }
            for cid, row in cross_shop_stats.iterrows()
        ]

        return {
            'targetShop': target_shop,
            'totalCustomers': int(len(target_customer_ids)),
            'newCustomers': int(len(new_ids)),
            'existingCustomers': int(len(existing_ids)),
            'newPct': round(len(new_ids) / len(target_customer_ids) * 100, 1) if len(target_customer_ids) > 0 else 0,
            'existingPct': round(len(existing_ids) / len(target_customer_ids) * 100, 1) if len(target_customer_ids) > 0 else 0,

            'revenueNew': round(rev_new, 2),
            'revenueExisting': round(rev_existing, 2),
            'revenueTotal': round(total_rev, 2),
            'revenueNewPct': round(rev_new / total_rev * 100, 1) if total_rev > 0 else 0,
            'revenueExistingPct': round(rev_existing / total_rev * 100, 1) if total_rev > 0 else 0,

            'avgSpendNew': round(rev_new / len(new_ids), 2) if len(new_ids) > 0 else 0,
            'avgSpendExisting': round(rev_existing / len(existing_ids), 2) if len(existing_ids) > 0 else 0,
            'avgSpendOverall': round(total_rev / len(target_customer_ids), 2) if len(target_customer_ids) > 0 else 0,

            'repeatRateNew': round(new_repeat_count / len(new_ids) * 100, 1) if len(new_ids) > 0 else 0,
            'repeatRateExisting': round(existing_repeat_count / len(existing_ids) * 100, 1) if len(existing_ids) > 0 else 0,
            'repeatRateOverall': overall_repeat_pct,

            'sourceShops': source_shops,
            'sourceRegions': source_regions,
            'internalLoyaltyDetails': internal_loyalty_details,
            'crossLoyaltyDetails': cross_loyalty_details,
            'monthlyTrends': calculate_monthly_loyalty_trends(df, target_shop)
        }
    except Exception as e:
        import traceback
        print(f"[ERROR] In loyalty analysis for {target_shop}: {e}")
        print(traceback.format_exc())
        return {'error': str(e), 'totalCustomers': 0}

def calculate_monthly_loyalty_trends(df, target_shop):
    """Calculate the monthly evolution of loyalty categories for a specific shop using vectorized grouping."""
    try:
        if df.empty:
            return []
            
        # 1. Filter to customers who have EVER visited the target shop
        target_customer_ids = df[df['Shop'] == target_shop]['Customer_ID'].unique()
        if len(target_customer_ids) == 0:
            return []
            
        target_cust_history = df[df['Customer_ID'].isin(target_customer_ids)].copy()
        target_cust_history['YearMonth'] = target_cust_history['Date'].dt.to_period('M')
        
        # 2. Daily categorization per customer (Target-Only vs Cross-Shop)
        # Find which shops each customer visited on ANY day
        cust_profile = target_cust_history.groupby(['YearMonth', 'Customer_ID'])['Shop'].unique()
        
        # Categorize
        # Note: We use the *entire* history to define if someone is Cross-Shop, 
        # but the user might want "Cross-Shop in that month". 
        # Let's go with: if they visited OTHER shops in the same month, they are Cross-Shoppers.
        is_cross = cust_profile.apply(lambda shops: len(shops) > 1 or (len(shops) == 1 and shops[0] != target_shop))
        
        # 3. Aggregate by month
        monthly_stats = is_cross.groupby('YearMonth').agg(['count', 'sum']).reset_index()
        monthly_stats.columns = ['YearMonth', 'total', 'crossShop']
        monthly_stats['targetOnly'] = monthly_stats['total'] - monthly_stats['crossShop']
        
        # 3b. Find the actual start month for this shop to filter out pre-opening months
        shop_transactions = df[df['Shop'] == target_shop]
        if shop_transactions.empty:
            return []
        first_month = shop_transactions['Date'].dt.to_period('M').min()
        monthly_stats = monthly_stats[monthly_stats['YearMonth'] >= first_month]
        
        # 4. Format
        results = []
        for _, row in monthly_stats.sort_values('YearMonth').iterrows():
            total = int(row['total'])
            cross = int(row['crossShop'])
            target = int(row['targetOnly'])
            results.append({
                'month': str(row['YearMonth']),
                'total': total,
                'targetOnly': target,
                'targetOnlyPct': round((target/total*100), 1) if total > 0 else 0,
                'crossShop': cross,
                'crossShopPct': round((cross/total*100), 1) if total > 0 else 0
            })
            
        # Return only the last 12 months for visual clarity
        return results[-12:]
    except Exception as e:
        print(f"[ERROR] calculate_monthly_loyalty_trends for {target_shop}: {e}")
        return []

        return {'error': str(e)}

def calculate_monthly_shop_overview(df):
    """Calculate monthly customer and units sold overview by shop using vectorized groupby."""
    try:
        empty_res = {
            'customers': {'shops': {}, 'totals': {}, 'months': []},
            'units': {'shops': {}, 'totals': {}, 'months': []},
            'products': {'months': [], 'products': {}},
            'debug': {'error': 'No data or Shop column missing'}
        }
        if df.empty or 'Shop' not in df.columns:
            return empty_res
        
        df_copy = df.dropna(subset=['Date']).copy()
        df_copy['YearMonth'] = df_copy['Date'].dt.to_period('M').astype(str)
        
        # Filter to relevant shops
        df_copy = df_copy[df_copy['Shop'].isin(SHOP_REGION_MAP.keys())]
        if df_copy.empty:
            return empty_res

        shops = sorted(df_copy['Shop'].unique())
        months = sorted(df_copy['YearMonth'].unique())
        
        # 1. Calculate Customers and Units per (Shop, Month) in one go
        grouped = df_copy.groupby(['Shop', 'YearMonth']).agg(
            unique_customers=('Customer_ID', 'nunique'),
            units_sold=('Customer_ID', 'count')
        ).unstack(fill_value=0)
        
        customer_data = grouped['unique_customers'].to_dict(orient='index')
        units_data = grouped['units_sold'].to_dict(orient='index')
        
        # Calculate totals per month
        monthly_grouped = df_copy.groupby('YearMonth').agg(
            unique_customers=('Customer_ID', 'nunique'),
            units_sold=('Customer_ID', 'count')
        )
        customer_totals = monthly_grouped['unique_customers'].to_dict()
        units_totals = monthly_grouped['units_sold'].to_dict()
        
        # 2. Product-wise Monthly Overview (Top 20 products)
        product_col = None
        for col in ['Product', 'Item', 'Product Name', 'Item Name', 'Items', 'Description', 'Specifics', 'Product_Name', 'ProductName', 'Items Purchased']:
            if col in df_copy.columns: product_col = col; break
            
        product_overview = {'months': months, 'products': {}}
        if product_col:
            # Get top 20 products by revenue overall
            top_products = df_copy.groupby(product_col)['Total'].sum().sort_values(ascending=False).head(20).index.tolist()
            
            # Map products to months
            prod_monthly = df_copy[df_copy[product_col].isin(top_products)].groupby([product_col, 'YearMonth'])['Customer_ID'].count().unstack(fill_value=0)
            product_overview['products'] = prod_monthly.to_dict(orient='index')
            
        return {
            'customers': {'shops': customer_data, 'totals': customer_totals, 'months': months},
            'units': {'shops': units_data, 'totals': units_totals, 'months': months},
            'products': product_overview
        }
    except Exception as e:
        print(f"[ERROR] Monthly Shop Overview: {str(e)}")
        return empty_res

def calculate_inactive_customers(df, days_threshold=30, last_month=None, last_year=None, shop_filter=None):
    """Identify customers who haven't made a purchase within the threshold days.
    Uses the latest date in the entire dataset as the reference point for 'today'.
    """
    try:
        if df.empty:
            return []
            
        # 1. Reference date is ALWAYS global max date
        reference_date = df['Date'].max()
        
        # 2. Get the VERY latest visit for EVERY customer across EVERY shop
        df_sorted = df.sort_values('Date')
        latest_ever = df_sorted.groupby('Customer_ID').tail(1).copy()
        
        # 3. Combine with stats
        cust_stats = df.groupby('Customer_ID').agg(
            totalSpend=('Total', 'sum'),
            totalVisits=('Date', 'nunique')
        ).reset_index()
        
        # Merge latest info
        # Note: latest_ever already has Customer_ID, Date, Shop, Product, First Name, Phone, Gender
        merged = latest_ever.merge(cust_stats, on='Customer_ID')
        
        # 4. Calculate inactivity period
        merged['daysInactive'] = (reference_date - merged['Date']).dt.days
        
        # 5. Apply filters
        # Shop Filter: Usually means "People whose LAST visit was at this shop"
        if shop_filter and shop_filter != 'all':
            merged = merged[merged['Shop'] == shop_filter]
            
        if last_month and last_year and last_month != 'all' and last_year != 'all':
            # Specific cohort: people whose last purchase was exactly in this month/year
            merged = merged[
                (merged['Date'].dt.month == int(last_month)) &
                (merged['Date'].dt.year == int(last_year))
            ]
        else:
            # Default: anyone who hasn't shopped for X days
            merged = merged[merged['daysInactive'] >= days_threshold]
            
        if merged.empty:
            return []
        
        # 6. Sort by spend
        merged = merged.sort_values(by='totalSpend', ascending=False)
        
        # 7. Format results
        results = []
        for _, row in merged.iterrows():
            results.append({
                'customerId': row['Customer_ID'],
                'firstName': str(row['First Name']),
                'phone': str(row['Phone']),
                'gender': str(row['Gender']),
                'totalSpend': round(float(row['totalSpend']), 2),
                'totalVisits': int(row['totalVisits']),
                'lastPurchaseDate': row['Date'].strftime('%Y-%m-%d'),
                'daysInactive': int(row['daysInactive']),
                'lastShop': str(row['Shop']),
                'lastProduct': str(row['Product'])
            })
            
        return results
    except Exception as e:
        import traceback
        print(f"[ERROR] calculate_inactive_customers: {str(e)}")
        print(traceback.format_exc())
        return []

def calculate_growth_rates(df):
    """Calculate full historical growth rates for customers for Monthly, Quarterly, and Yearly periods"""
    try:
        if df.empty or 'Customer_ID' not in df.columns:
            return {'monthly': [], 'quarterly': [], 'yearly': []}
            
        df_copy = df.copy()
        if 'Shop' in df_copy.columns:
            # Filter to only include shops in our map for consistency with other metrics
            df_copy = df_copy[df_copy['Shop'].isin(SHOP_REGION_MAP.keys())]
            
        if df_copy.empty:
            return {'monthly': [], 'quarterly': [], 'yearly': []}

        df_copy['Date'] = pd.to_datetime(df_copy['Date'])
        
        def compute_series(stats):
            series = []
            for i in range(len(stats)):
                curr_val = stats.iloc[i]['customers']
                prev_val = stats.iloc[i-1]['customers'] if i > 0 else 0
                
                growth = 0
                if i > 0 and prev_val > 0:
                    growth = round(((curr_val - prev_val) / prev_val * 100), 2)
                
                series.append({
                    'period': str(stats.index[i]),
                    'current': int(curr_val),
                    'previous': int(prev_val),
                    'growth': growth,
                    'is_start': i == 0
                })
            return series

        # 1. Monthly Series
        df_copy['Period_M'] = df_copy['Date'].dt.to_period('M')
        m_stats = df_copy.groupby('Period_M').agg({'Customer_ID': 'nunique'}).rename(columns={'Customer_ID': 'customers'}).sort_index()
        
        # 2. Quarterly Series
        df_copy['Period_Q'] = df_copy['Date'].dt.to_period('Q')
        q_stats = df_copy.groupby('Period_Q').agg({'Customer_ID': 'nunique'}).rename(columns={'Customer_ID': 'customers'}).sort_index()
        
        # 3. Yearly Series
        df_copy['Period_Y'] = df_copy['Date'].dt.to_period('Y')
        y_stats = df_copy.groupby('Period_Y').agg({'Customer_ID': 'nunique'}).rename(columns={'Customer_ID': 'customers'}).sort_index()

        return {
            'monthly': compute_series(m_stats),
            'quarterly': compute_series(q_stats),
            'yearly': compute_series(y_stats)
        }
        
    except Exception as e:
        print(f"[ERROR] Error in calculate_growth_rates: {str(e)}")
        return {'monthly': [], 'quarterly': [], 'yearly': []}
        
    except Exception as e:
        print(f"[ERROR] Error in calculate_growth_rates: {str(e)}")
        return []

@app.route('/')
def index():
    """Render the dashboard"""
    return render_template('index.html')

def calculate_visit_sequence_spend(df):
    """Optimized calculation of average spend Trajectory up to 5th visit."""
    try:
        if df.empty or 'Customer_ID' not in df.columns:
            return []
            
        # Optimization: Only select necessary columns
        sub_df = df[['Customer_ID', 'Date', 'Total']]

        # Aggregate daily spend per customer to define a "visit"
        daily = sub_df.groupby(['Customer_ID', sub_df['Date'].dt.date], sort=False)['Total'].sum().reset_index()
        
        # Sort values
        daily = daily.sort_values(['Customer_ID', 'Date'])
        
        # Assign visit numbers
        daily['vnum'] = daily.groupby('Customer_ID').cumcount() + 1
        
        # Filter to 5 visits early
        daily = daily[daily['vnum'] <= 5]
        
        # Calculate cumulative total spend per customer
        daily['cum_spend'] = daily.groupby('Customer_ID')['Total'].cumsum()

        # Stats per visit number - Average of the total spent up to that visit
        stats = daily.groupby('vnum').agg(
            avg_spend=('Total', 'mean'),
            avg_cum_spend=('cum_spend', 'mean'),
            cust_count=('Customer_ID', 'count')
        ).reset_index()
        
        results = []
        for _, row in stats.iterrows():
            v = int(row['vnum'])
            suffix = {1:'st', 2:'nd', 3:'rd'}.get(v, 'th')
            results.append({
                'visitOrder': f"{v}{suffix} Visit",
                'avgSpend': float(row['avg_spend']),
                'cumulativeAvgSpend': float(row['avg_cum_spend']),
                'customerCount': int(row['cust_count'])
            })
            
        return results
    except Exception as e:
        print(f"[ERROR] calculate_visit_sequence_spend: {str(e)}")
        return []

def _compute_global_results(df, loyalty_logic='cross-shop'):
    """Compute overall dashboard metrics only (no per-shop breakdown). Fast path."""
    global global_results_cache

    df = _prepare_working_df(df)

    overall_results = {
        'overall': calculate_overall_performance(df),
        'overallBreakdown': calculate_overall_repeat_breakdown(df),
        'overview': calculate_overview(df),
        'monthly': calculate_monthly_data(df),
        'weekly': calculate_weekly_data(df),
        'monthlyRepeatBreakdown': calculate_monthly_repeat_breakdown(df),
        'quarterly': calculate_quarterly_data(df),
        'semiAnnual': calculate_semiannual_performance(df),
        'semiAnnualBreakdown': calculate_semiannual_repeat_breakdown(df),
        'yearly': calculate_yearly_data(df),
        'regions': calculate_regional_data(df),
        'gender': calculate_gender_performance(df),
        'products': calculate_product_performance(df),
        'topShopsByRegion': calculate_top_shops_by_region(df),
        'cumulativeRetention': calculate_cumulative_retention(df),
        'visitIntervals': calculate_visit_interval_distribution(df),
        'advancedProducts': analyze_combos_and_affinity(df),
        'regionalProducts': calculate_regional_top_products(df),
        'monthlyShopOverview': calculate_monthly_shop_overview(df),
        'growthRates': calculate_growth_rates(df),
        'visitSequenceSpend': calculate_visit_sequence_spend(df),
    }

    monthly_data = overall_results.get('monthly', [])
    marketing_highlights = {
        'totalMarketingSpend': 0, 'totalNewCustomers': 0, 'overallCAC': 0,
        'bestAcquisitionMonth': None, 'roiLeaderMonth': None, 'topEfficiencyMonth': None
    }
    if monthly_data:
        total_spend = sum(m.get('marketingSpend', 0) for m in monthly_data)
        total_new   = sum(m.get('newCustomers', 0)  for m in monthly_data)
        marketing_highlights['totalMarketingSpend'] = round(total_spend, 2)
        marketing_highlights['totalNewCustomers']   = int(total_new)
        marketing_highlights['overallCAC'] = round(total_spend / total_new, 2) if total_new > 0 else 0
        best_acq = max(monthly_data, key=lambda x: x.get('newCustomers', 0), default=None)
        if best_acq:
            marketing_highlights['bestAcquisitionMonth'] = f"{best_acq['period']} ({int(best_acq['newCustomers'])} new)"
        roi_months = [m for m in monthly_data if m.get('marketingSpend', 0) > 0]
        if roi_months:
            best_roi = max(roi_months, key=lambda x: x.get('totalRevenue', 0) / x.get('marketingSpend', 1), default=None)
            if best_roi:
                roi_val = round(best_roi['totalRevenue'] / best_roi['marketingSpend'], 2)
                marketing_highlights['roiLeaderMonth'] = f"{best_roi['period']} ({roi_val}x ROI)"
            best_eff = min(roi_months, key=lambda x: x.get('cac', 999999), default=None)
            if best_eff:
                marketing_highlights['topEfficiencyMonth'] = f"{best_eff['period']} (CAC: {best_eff['cac']})"

    data = {'marketingHighlights': marketing_highlights, 'shops': {}}
    data.update(overall_results)
    global_results_cache = data
    return data


def _compute_shop_results(df, loyalty_logic='cross-shop'):
    """Compute per-shop metrics. Slow path — called from /api/shops."""
    global shops_results_cache

    df = _prepare_working_df(df)
    priority_shops = ['Ktda', 'Kisii', 'Busia', 'Rongai']
    shops = {}

    if 'Shop' not in df.columns:
        return {}

    available_shops = [s for s in df['Shop'].unique() if s in SHOP_REGION_MAP]

    def process_shop_data(shop):
        try:
            shop_df = df[df['Shop'] == shop]
            if shop_df.empty:
                return shop, None
            is_priority = any(p.lower() == shop.lower() for p in priority_shops)
            shop_results = {
                'overall': calculate_overall_performance(shop_df),
                'overallBreakdown': calculate_overall_repeat_breakdown(shop_df),
                'overview': calculate_overview(shop_df),
                'monthly': calculate_monthly_data(shop_df),
                'weekly': calculate_weekly_data(shop_df),
                'monthlyRepeatBreakdown': calculate_monthly_repeat_breakdown(shop_df),
                'quarterly': calculate_quarterly_data(shop_df),
                'semiAnnual': calculate_semiannual_performance(shop_df),
                'semiAnnualBreakdown': calculate_semiannual_repeat_breakdown(shop_df),
                'yearly': calculate_yearly_data(shop_df),
                'visitIntervals': calculate_visit_interval_distribution(shop_df),
                'growthRates': calculate_growth_rates(shop_df),
            }
            if is_priority:
                shop_results['loyaltyAnalysis'] = calculate_shop_loyalty_analysis(df, shop, logic=loyalty_logic)
            return shop, shop_results
        except Exception as e:
            print(f"[ERROR] processing shop {shop}: {e}")
            return shop, None

    with ThreadPoolExecutor(max_workers=os.cpu_count() or 4) as executor:
        futures = {executor.submit(process_shop_data, s): s for s in available_shops}
        for future in concurrent.futures.as_completed(futures):
            shop, result = future.result()
            if result:
                shops[shop] = result

    # Build priority-shop loyalty keys so existing frontend keys still work
    for p_shop in priority_shops:
        actual = next((s for s in shops if s.lower() == p_shop.lower()), None)
        if actual and 'loyaltyAnalysis' in shops[actual]:
            la = shops[actual]['loyaltyAnalysis']
            shops[f'__{p_shop.lower()}AnalysisCross']    = la if loyalty_logic == 'cross-shop' else calculate_shop_loyalty_analysis(df, actual, logic='cross-shop')
            shops[f'__{p_shop.lower()}AnalysisInternal'] = la if loyalty_logic == 'internal'   else calculate_shop_loyalty_analysis(df, actual, logic='internal')
        else:
            shops[f'__{p_shop.lower()}AnalysisCross']    = calculate_shop_loyalty_analysis(df, p_shop, logic='cross-shop')
            shops[f'__{p_shop.lower()}AnalysisInternal'] = calculate_shop_loyalty_analysis(df, p_shop, logic='internal')

    shops_results_cache = shops
    return shops


def _compute_all_results(df, loyalty_logic='cross-shop'):
    """Legacy: compute everything at once (used for filtered requests)."""
    global computed_results_cache
    data  = _compute_global_results(df, loyalty_logic)
    shops = _compute_shop_results(df, loyalty_logic)
    # Merge loyalty analysis keys from shops into top-level data
    for key in list(shops.keys()):
        if key.startswith('__'):
            data[key[2:]] = shops.pop(key)
    data['shops'] = shops
    computed_results_cache = data
    return data

@app.route('/api/data')
def get_data():
    """API endpoint to get all dashboard data"""
    global computed_results_cache
    
    # Get filters from query parameters
    filter_year = request.args.get('year')
    filter_month = request.args.get('month')
    filter_quarter = request.args.get('quarter')
    filter_half = request.args.get('half')
    filter_week = request.args.get('week')
    loyalty_logic = request.args.get('loyaltyMode', 'cross-shop')
    
    # Check if any filter is actually applied (not 'all' and not None)
    is_filtered = any([
        filter_year and filter_year != 'all',
        filter_month and filter_month != 'all',
        filter_quarter and filter_quarter != 'all',
        filter_half and filter_half != 'all',
        filter_week and filter_week != 'all'
    ])
    
    try:
        # ── Fast path: serve precomputed analytics from Supabase cache ─────────
        if not is_filtered and SUPABASE_KEY:
            _ac = _read_analytics_cache()
            if _ac:
                _res = _ac['result']
                _res['cache_status'] = {
                    'last_updated': _ac.get('updated_at', 'unknown'),
                    'type': 'analytics_cache',
                }
                return jsonify(_res)

        # 1. Fetch data
        df = get_customer_data()

        # 2. Apply time filters if provided
        if is_filtered:
            print(f"[INFO] Applying time filters: year={filter_year}, month={filter_month}, quarter={filter_quarter}, week={filter_week}")
            df_work = df.copy()
            
            if filter_year and filter_year != 'all':
                df_work = df_work[df_work['Date'].dt.year == int(filter_year)]
            
            if filter_month and filter_month != 'all':
                df_work = df_work[df_work['Date'].dt.month == int(filter_month)]
                
            if filter_quarter and filter_quarter != 'all':
                # Remove 'Q' if present (e.g., 'Q1' -> 1)
                q_val = filter_quarter.replace('Q', '')
                df_work = df_work[df_work['Date'].dt.quarter == int(q_val)]
            
            if filter_half and filter_half != 'all':
                if filter_half == 'H1':
                    df_work = df_work[df_work['Date'].dt.month <= 6]
                else:
                    df_work = df_work[df_work['Date'].dt.month > 6]
            
            if filter_week and filter_week != 'all':
                # Week format is "W##" (e.g., "W01" for week 1)
                # Add YearWeek column if not present
                df_work['YearWeek'] = df_work['Date'].dt.isocalendar().year.astype(str) + '-W' + df_work['Date'].dt.isocalendar().week.astype(str).str.zfill(2)
                week_key = f"{filter_year}-W{filter_week.zfill(2)}"
                df_work = df_work[df_work['YearWeek'] == week_key]
                
            df = df_work
        
        # 3. Fast path: return global-only cache for unfiltered requests
        if not is_filtered and global_results_cache is not None:
            global_results_cache['cache_status'] = {
                'last_updated': datetime.fromtimestamp(last_fetch_time).strftime('%Y-%m-%d %H:%M:%S') if last_fetch_time else "Unknown",
                'type': 'computed_memory'
            }
            return jsonify(global_results_cache)

        print(f"[INFO] Calculating global metrics (Logic: {loyalty_logic})...")

        # Unfiltered: compute global only (shops are loaded separately via /api/shops)
        if not is_filtered:
            result = _compute_global_results(df, loyalty_logic=loyalty_logic)
        else:
            # Filtered views need shop data for the shop-selector view
            result = _compute_all_results(df, loyalty_logic=loyalty_logic)

        # SPECIAL: Inject CAC for the current filtered period if it's a single month
        if filter_month and filter_month != 'all' and filter_year and filter_year != 'all':
            month_key = f"{filter_year}-{int(filter_month):02d}"
            full_df = get_customer_data()
            full_monthly = calculate_monthly_data(full_df)
            matching = next((m for m in full_monthly if m['period'] == month_key), None)
            if matching and matching.get('marketingSpend', 0) > 0:
                if 'overview' in result:
                    result['overview']['cac'] = matching['cac']
                    result['overview']['marketingSpend'] = matching['marketingSpend']
                if 'overall' in result:
                    result['overall']['cac'] = matching['cac']
                    result['overall']['marketingSpend'] = matching['marketingSpend']
        
        # SPECIAL: Inject CAC for the current filtered period if it's a single week
        if filter_week and filter_week != 'all' and filter_year and filter_year != 'all':
            week_key = f"{filter_year}-W{filter_week.zfill(2)}"
            full_df = get_customer_data()
            full_weekly = calculate_weekly_data(full_df)
            matching = next((w for w in full_weekly if w['period'] == week_key), None)
            if matching and matching.get('marketingSpend', 0) > 0:
                if 'overview' in result:
                    result['overview']['cac'] = matching['cac']
                    result['overview']['marketingSpend'] = matching['marketingSpend']
                if 'overall' in result:
                    result['overall']['cac'] = matching['cac']
                    result['overall']['marketingSpend'] = matching['marketingSpend']
        
        result['cache_status'] = {
            'last_updated': datetime.fromtimestamp(last_fetch_time).strftime('%Y-%m-%d %H:%M:%S') if last_fetch_time else "Unknown",
            'type': 'computed_fresh'
        }

        # Persist unfiltered result so Vercel's next cold-start serves it instantly
        if not is_filtered:
            _write_analytics_cache(result)

        return jsonify(result)
    
    except Exception as e:
        import traceback
        print(f"[ERROR] in /api/data: {str(e)}")
        print(f"[TRACEBACK] {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500



@app.route('/api/shops')
def get_shops():
    """Return per-shop metrics. Loaded asynchronously by the frontend after initial render."""
    global shops_results_cache
    try:
        loyalty_logic = request.args.get('loyaltyMode', 'cross-shop')
        if shops_results_cache is not None:
            return jsonify(shops_results_cache)
        print("[INFO] Computing shop metrics...")
        df = get_customer_data()
        shops = _compute_shop_results(df, loyalty_logic=loyalty_logic)
        # Expose priority-shop loyalty keys at top level for backward compatibility
        response = {}
        for key in list(shops.keys()):
            if key.startswith('__'):
                response[key[2:]] = shops[key]
            else:
                response[key] = shops[key]
        return jsonify(response)
    except Exception as e:
        import traceback
        print(f"[ERROR] in /api/shops: {str(e)}")
        print(f"[TRACEBACK] {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ktda-customer-analysis')
def get_ktda_customer_analysis():
    """API endpoint to get KTDA customer classification analysis"""
    try:
        df = get_customer_data()
        analysis = calculate_shop_loyalty_analysis(df, 'Ktda')
        return jsonify(analysis)
    except Exception as e:
        import traceback
        print(f"[ERROR] in /api/ktda-customer-analysis: {str(e)}")
        print(f"[TRACEBACK] {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/inactive-customers')
def get_inactive_customers():
    """API endpoint for inactive customers list with optional month/year cohort and shop filter"""
    try:
        df = get_customer_data()
        if df is None:
            return jsonify([])
            
        days = request.args.get('days', 30, type=int)
        month = request.args.get('month', None)
        year = request.args.get('year', None)
        shop = request.args.get('shop', None)
        
        inactive = calculate_inactive_customers(df, days, month, year, shop)
        return jsonify(inactive)
    except Exception as e:
        print(f"[ERROR] API inactive-customers: {str(e)}")
        return jsonify([])

@app.route('/api/export/inactive-customers')
def export_inactive_customers():
    """Export inactive customers to CSV with optional cohort and shop filter"""
    try:
        from flask import Response
        df = get_customer_data()
        if df is None:
            return "No data found", 404
            
        days = request.args.get('days', 30, type=int)
        month = request.args.get('month', None)
        year = request.args.get('year', None)
        shop = request.args.get('shop', None)
        
        inactive_list = calculate_inactive_customers(df, days, month, year, shop)
        
        if not inactive_list:
            return "No inactive customers found", 404
        
        export_df = pd.DataFrame(inactive_list)
        # Rename for export
        export_df = export_df.rename(columns={
            'firstName': 'First Name',
            'phone': 'Phone',
            'gender': 'Gender',
            'lastPurchaseDate': 'Last Purchase',
            'daysInactive': 'Days Inactive',
            'totalSpend': 'Total Spend',
            'totalVisits': 'Total Visits',
            'lastShop': 'Last Shop',
            'lastProduct': 'Last Bag Bought'
        })
        
        # Reorder
        cols = ['First Name', 'Phone', 'Gender', 'Last Purchase', 'Days Inactive', 'Total Spend', 'Total Visits', 'Last Shop', 'Last Bag Bought']
        export_df = export_df[cols]
        
        csv_output = export_df.to_csv(index=False)
        filename = f"inactive_customers_{days}_days.csv"
        
        return Response(csv_output, mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}', 'Content-Type': 'text/csv; charset=utf-8'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/refresh-now', methods=['POST'])
def refresh_now():
    """Drop every cache and re-read from Supabase."""
    global cached_data, last_fetch_time, computed_results_cache, global_results_cache, shops_results_cache
    cached_data = None
    last_fetch_time = None
    computed_results_cache = None
    global_results_cache   = None
    shops_results_cache    = None
    
    # Remove persistent cache file to force fresh fetch
    if os.path.exists(CACHE_FILE):
        try:
            os.remove(CACHE_FILE)
            print(f"[INFO] Removed persistent cache file: {CACHE_FILE}")
        except Exception as e:
            print(f"[WARNING] Could not remove persistent cache: {e}")
            
    return get_data()

@app.route('/api/export/repeat-customers')
def export_repeat_customers():
    """Export all transaction rows for retained (repeat) customers, including products.
    Optional query param: ?shop=<ShopName>
    """
    try:
        from flask import Response
        df = get_customer_data()

        # Optional shop filter
        shop_filter = request.args.get('shop', '').strip()
        if shop_filter and shop_filter in df['Shop'].values:
            df_work = df[df['Shop'] == shop_filter].copy()
            filename = f"repeat_customers_{shop_filter.lower().replace(' ', '_')}.csv"
        else:
            df_work = df.copy()
            filename = "repeat_customers_all.csv"

        # Identify repeat IDs (visited on at least 2 different calendar days)
        df_work['Visit_Date'] = df_work['Date'].dt.date
        visit_days = df_work.groupby('Customer_ID')['Visit_Date'].nunique()
        repeat_ids = visit_days[visit_days >= 2].index

        # Filter for the target customers
        # IMPORTANT: We use the full original row set for these customers
        export_df = df_work[df_work['Customer_ID'].isin(repeat_ids)].copy()

        if export_df.empty:
            return Response("First Name,Phone,Gender,Shop,Date,Product,Price,Quantity,Total,Total_Customer_Spend\n",
                            mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={filename}'})

        # Calculate total spend per customer for sorting
        spend_map = export_df.groupby('Customer_ID')['Total'].sum().to_dict()
        export_df['Total_Customer_Spend'] = export_df['Customer_ID'].map(spend_map)

        # Select and reorder columns for the call center
        cols = ['First Name', 'Phone', 'Gender', 'Shop', 'Date', 'Product', 'Price', 'Quantity', 'Total', 'Total_Customer_Spend', 'Customer_ID']
        # Check matching columns
        existing_cols = [c for c in cols if c in export_df.columns]
        export_df = export_df[existing_cols]

        # Sort by Customer Spend (Descending) then Date
        export_df = export_df.sort_values(by=['Total_Customer_Spend', 'Customer_ID', 'Date'], ascending=[False, True, False])
        
        # Now drop Customer_ID if we don't want it in the final CSV
        if 'Customer_ID' in export_df.columns:
            export_df = export_df.drop(columns=['Customer_ID'])

        csv_output = export_df.to_csv(index=False)
        print(f"[INFO] Exporting {len(export_df)} transaction rows for {len(repeat_ids)} repeat customers")
        
        return Response(csv_output, mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}', 'Content-Type': 'text/csv; charset=utf-8'})

    except Exception as e:
        import traceback
        print(f"[ERROR] export_repeat_customers: {str(e)}")
        print(f"[TRACEBACK] {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/one-time-customers')
def export_one_time_customers():
    """Export all transaction rows for one-time (single-visit) customers, including products.
    Optional query param: ?shop=<ShopName>
    """
    try:
        from flask import Response
        df = get_customer_data()

        # Optional shop filter
        shop_filter = request.args.get('shop', '').strip()
        if shop_filter and shop_filter in df['Shop'].values:
            df_work = df[df['Shop'] == shop_filter].copy()
            filename = f"one_time_customers_{shop_filter.lower().replace(' ', '_')}.csv"
        else:
            df_work = df.copy()
            filename = "one_time_customers_all.csv"

        # Identify one-time IDs (visited on exactly 1 unique calendar day)
        df_work['Visit_Date'] = df_work['Date'].dt.date
        visit_days = df_work.groupby('Customer_ID')['Visit_Date'].nunique()
        one_time_ids = visit_days[visit_days == 1].index

        # Filter for the target customers
        export_df = df_work[df_work['Customer_ID'].isin(one_time_ids)].copy()

        if export_df.empty:
            return Response("First Name,Phone,Gender,Shop,Date,Product,Price,Quantity,Total,Total_Customer_Spend\n",
                            mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={filename}'})

        # Calculate total spend per customer for sorting
        spend_map = export_df.groupby('Customer_ID')['Total'].sum().to_dict()
        export_df['Total_Customer_Spend'] = export_df['Customer_ID'].map(spend_map)

        # Select relevant columns
        cols = ['First Name', 'Phone', 'Gender', 'Shop', 'Date', 'Product', 'Price', 'Quantity', 'Total', 'Total_Customer_Spend', 'Customer_ID']
        existing_cols = [c for c in cols if c in export_df.columns]
        export_df = export_df[existing_cols]

        # Sort by Customer Spend (Descending)
        export_df = export_df.sort_values(by=['Total_Customer_Spend', 'Customer_ID'], ascending=[False, True])
        
        # Drop Customer_ID for final output
        if 'Customer_ID' in export_df.columns:
            export_df = export_df.drop(columns=['Customer_ID'])

        csv_output = export_df.to_csv(index=False)
        print(f"[INFO] Exporting {len(export_df)} transaction rows for {len(one_time_ids)} one-time customers")
        
        return Response(csv_output, mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}', 'Content-Type': 'text/csv; charset=utf-8'})

    except Exception as e:
        import traceback
        print(f"[ERROR] export_one_time_customers: {str(e)}")
        print(f"[TRACEBACK] {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

def _refresh_analytics(df=None):
    """Recompute the dashboard analytics and store them in Supabase.

    The dashboard must never bulk-read `sales` on a page load — free-tier
    Supabase cannot serve a 250k-row read reliably. Everything that changes the
    data calls this once, and every page load then reads a single small row.
    """
    global cached_data, last_fetch_time, computed_results_cache
    global global_results_cache, shops_results_cache

    if df is None:
        df = _load_from_supabase()

    cached_data            = df
    last_fetch_time        = time.time()
    computed_results_cache = None
    shops_results_cache    = None
    _save_working_copy(df)

    results = _compute_global_results(df)
    results['cache_status'] = {
        'last_updated': datetime.fromtimestamp(last_fetch_time).strftime('%Y-%m-%d %H:%M:%S'),
        'type': 'analytics_cache',
    }
    _write_analytics_cache(results)
    print(f"[INFO] Analytics refreshed over {len(df)} records")
    return len(df)


@app.route('/api/upload', methods=['POST'])
def upload_data():
    """Append CSV rows straight into Supabase, then refresh the analytics."""
    if not DATABASE_URL:
        return jsonify({'error': 'DATABASE_URL env var not set'}), 500

    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        name = file.filename.lower()
        if name.endswith('.csv'):
            df_up = pd.read_csv(io.StringIO(file.read().decode('utf-8-sig')))
        elif name.endswith(('.xlsx', '.xls')):
            # Phone must stay a string — Excel stores it as a number and would
            # otherwise turn 0712345678 into 712345678.0 and break customer
            # matching, since repeat visits are keyed on the phone number.
            df_up = pd.read_excel(file, dtype={'Phone': str})
        else:
            return jsonify({'error': 'Upload a .csv or .xlsx file'}), 400

        # Accept the same column aliases the old Sheets import did.
        if 'Shop' not in df_up.columns and 'Location' in df_up.columns:
            df_up.rename(columns={'Location': 'Shop'}, inplace=True)
        if 'Gender' not in df_up.columns and 'Female' in df_up.columns:
            df_up.rename(columns={'Female': 'Gender'}, inplace=True)

        required = ['Date', 'First Name', 'Phone', 'Price', 'Shop']
        missing = [c for c in required if c not in df_up.columns]
        if missing:
            return jsonify({
                'error': f'Missing required columns: {", ".join(missing)}'
            }), 400

        _ensure_supabase_tables()

        # Same cleaning the dashboard applies to everything else, so uploaded
        # rows are indistinguishable from rows already in the table.
        df_up['Date'] = pd.to_datetime(df_up['Date'], errors='coerce')
        bad_dates = int(df_up['Date'].isna().sum())
        df_up = df_up.dropna(subset=['Date']).copy()
        if df_up.empty:
            return jsonify({'error': 'No rows with a valid Date were found'}), 400

        if 'Shop' in df_up.columns:
            df_up['Shop'] = df_up['Shop'].astype(str).str.strip().str.title()
        if 'Gender' in df_up.columns:
            df_up = df_up[
                df_up['Gender'].astype(str).str.lower().str.strip() != 'organization'
            ].copy()

        for col in ('Price', 'Quantity', 'Total', 'MARKETING EXPENSE'):
            if col in df_up.columns:
                df_up[col] = pd.to_numeric(
                    df_up[col].astype(str).str.replace(r'[^\d.]', '', regex=True),
                    errors='coerce')
        if 'Quantity' not in df_up.columns:
            df_up['Quantity'] = 1
        df_up['Quantity'] = df_up['Quantity'].fillna(1).replace(0, 1)
        if 'Total' not in df_up.columns:
            df_up['Total'] = df_up['Price'] * df_up['Quantity']
        df_up['Total'] = df_up['Total'].fillna(df_up['Price'] * df_up['Quantity'])

        keep   = [c for c in _APP_TO_SB if c in df_up.columns]
        df_out = df_up[keep].copy()
        df_out['Date'] = df_out['Date'].dt.strftime('%Y-%m-%d')
        df_out.rename(columns=_APP_TO_SB, inplace=True)

        n_added = _push_to_supabase_sql(df_out, truncate=False)

        # Recompute from the working set we already have plus the new rows.
        # Re-reading the whole table back out of Supabase just to add a handful
        # of rows is the slow, failure-prone path — avoid it when we can.
        total = None
        try:
            df_prev = get_customer_data()
            if df_prev is not None and not df_prev.empty:
                df_new = df_up.copy()
                df_new['Customer_ID'] = _customer_id(df_new['Phone'])
                total = _refresh_analytics(
                    pd.concat([df_prev, df_new], ignore_index=True))
        except Exception as e:
            print(f"[WARNING] Could not extend the working set ({e}); "
                  f"falling back to a full reload")
        if total is None:
            total = _refresh_analytics()

        msg = f'Added {n_added} record(s). Dashboard now covers {total} records.'
        if bad_dates:
            msg += f' ({bad_dates} row(s) skipped — unreadable Date.)'
        return jsonify({'success': True, 'message': msg,
                        'records_uploaded': n_added,
                        'rows_skipped': bad_dates,
                        'total_records': total})

    except Exception as e:
        import traceback
        print(f"[ERROR] Upload failed: {traceback.format_exc()}")
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500


@app.route('/api/recompute', methods=['POST'])
def recompute_analytics():
    """Rebuild the cached analytics from whatever is currently in Supabase.

    Use after editing rows directly in the Supabase table editor.
    """
    if not DATABASE_URL:
        return jsonify({'error': 'DATABASE_URL env var not set'}), 500
    try:
        _ensure_supabase_tables()
        total = _refresh_analytics()
        return jsonify({'success': True, 'total_records': total,
                        'message': f'Analytics rebuilt over {total} records.'})
    except Exception as e:
        import traceback
        print(f"[ERROR] Recompute failed: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500



def _location_month_options(df, shop=None):
    """Months that actually have data, newest first, as [{value,label}]."""
    d = df.dropna(subset=['Date'])
    if shop and str(shop).lower() != 'all':
        d = d[d['Shop'].astype(str).str.strip().str.lower() == str(shop).strip().lower()]
    if d.empty:
        return []
    keys = sorted(d['Date'].dt.strftime('%Y-%m').unique(), reverse=True)
    return [{'value': k,
             'label': pd.Period(k, freq='M').strftime('%B %Y')} for k in keys]


def _calculate_location_metrics(df, shop, month, scope='shop',
                                top_n=10, best_n=50):
    """Per-location, per-month metrics: spend, fast movers, best customers,
    retention.

    Mirrors the four sheets of the KTDA metrics workbook so a location report
    can be read straight off the dashboard instead of being rebuilt by hand.

    `scope` decides what counts as a returning customer:
      'shop' — seen at THIS location before the month (a location's own history)
      'all'  — seen at ANY location before the month (first-ever visit)
    """
    df = df.dropna(subset=['Date']).copy()

    shop_key = str(shop).strip().lower()
    shop_df = df[df['Shop'].astype(str).str.strip().str.lower() == shop_key]
    if shop_df.empty:
        return None

    period = pd.Period(str(month), freq='M')
    start, end = period.start_time, period.end_time
    m = shop_df[(shop_df['Date'] >= start) & (shop_df['Date'] <= end)].copy()
    if m.empty:
        return None

    has_product = 'Product' in m.columns
    has_gender = 'Gender' in m.columns
    if 'Quantity' not in m.columns:
        m['Quantity'] = 1
    m['Quantity'] = pd.to_numeric(m['Quantity'], errors='coerce').fillna(1)
    m['Total'] = pd.to_numeric(m['Total'], errors='coerce').fillna(0)

    def money(x):
        return round(float(x), 2)

    # ── 1. Spend metrics ─────────────────────────────────────────────────────
    per_cust = m.groupby('Customer_ID')['Total'].sum()
    revenue = m['Total'].sum()
    spend = {
        'totalTransactions': int(len(m)),
        'uniqueCustomers': int(m['Customer_ID'].nunique()),
        'totalRevenue': money(revenue),
        'avgPerTransaction': money(m['Total'].mean()),
        'minTransaction': money(m['Total'].min()),
        'maxTransaction': money(m['Total'].max()),
        'avgPerCustomer': money(per_cust.mean()),
        'medianPerCustomer': money(per_cust.median()),
        'highestCustomerSpend': money(per_cust.max()),
        'lowestCustomerSpend': money(per_cust.min()),
    }

    by_product = []
    if has_product:
        g = m.groupby('Product').agg(
            units=('Quantity', 'sum'),
            transactions=('Total', 'size'),
            revenue=('Total', 'sum'),
        ).reset_index().sort_values('revenue', ascending=False)
        for _, r in g.iterrows():
            trx = int(r['transactions'])
            by_product.append({
                'product': str(r['Product']),
                'units': int(r['units']),
                'transactions': trx,
                'revenue': money(r['revenue']),
                'avgPerTransaction': money(r['revenue'] / trx) if trx else 0,
            })

    # ── 2. Fast moving products ──────────────────────────────────────────────
    fast_movers, daily_trend = [], {'days': [], 'series': []}
    if has_product:
        units = m.groupby('Product')['Quantity'].sum().sort_values(ascending=False)
        total_units = float(units.sum())
        top = units.head(top_n)
        # Two denominators, because they answer different questions and the
        # source workbook used the narrower one: share of everything sold that
        # month, vs share within the top-10 group itself.
        top_units = float(top.sum())
        for rank, (prod, u) in enumerate(top.items(), start=1):
            fast_movers.append({
                'rank': rank,
                'product': str(prod),
                'units': int(u),
                'pctOfUnits': round(float(u) / total_units * 100, 1) if total_units else 0,
                'pctOfTopUnits': round(float(u) / top_units * 100, 1) if top_units else 0,
            })

        # Day-of-month x top product matrix, so a spike can be traced to a date.
        top_names = list(top.index)
        dm = m[m['Product'].isin(top_names)].copy()
        dm['Day'] = dm['Date'].dt.day
        days = list(range(1, int(period.days_in_month) + 1))
        pivot = dm.pivot_table(index='Day', columns='Product', values='Quantity',
                               aggfunc='sum', fill_value=0)
        pivot = pivot.reindex(index=days, fill_value=0)
        daily_trend = {
            'days': days,
            'series': [{'product': str(p),
                        'data': [int(pivot[p].get(d, 0)) for d in days]}
                       for p in top_names if p in pivot.columns],
        }

    # ── 3. Best customers ────────────────────────────────────────────────────
    agg = {
        'totalSpend': ('Total', 'sum'),
        'transactions': ('Total', 'size'),
        'visitDays': ('Date', lambda s: s.dt.normalize().nunique()),
        'lastVisit': ('Date', 'max'),
    }
    if 'First Name' in m.columns:
        agg['name'] = ('First Name', 'first')
    if has_gender:
        agg['gender'] = ('Gender', 'first')
    if 'Phone' in m.columns:
        agg['phone'] = ('Phone', 'first')

    b = m.groupby('Customer_ID').agg(**agg).reset_index()
    b = b.sort_values('totalSpend', ascending=False).head(best_n)

    prods_by_cust = {}
    if has_product:
        prods_by_cust = (m[m['Customer_ID'].isin(b['Customer_ID'])]
                         .groupby('Customer_ID')['Product']
                         .apply(lambda s: sorted({str(x) for x in s if str(x).strip()}))
                         .to_dict())

    best_customers = []
    for rank, (_, r) in enumerate(b.iterrows(), start=1):
        trx = int(r['transactions'])
        best_customers.append({
            'rank': rank,
            'name': str(r['name']) if 'name' in b.columns and pd.notna(r.get('name')) else '',
            'phone': _customer_id(pd.Series([r.get('phone', r['Customer_ID'])])).iloc[0],
            'gender': str(r['gender']) if 'gender' in b.columns and pd.notna(r.get('gender')) else '',
            'totalSpend': money(r['totalSpend']),
            'transactions': trx,
            'visitDays': int(r['visitDays']),
            'avgPerTransaction': money(r['totalSpend'] / trx) if trx else 0,
            'lastVisit': r['lastVisit'].strftime('%d %b %Y'),
            'products': prods_by_cust.get(r['Customer_ID'], []),
        })

    # ── 4. Retention ─────────────────────────────────────────────────────────
    history = shop_df if scope == 'shop' else df
    prior = set(history[history['Date'] < start]['Customer_ID'].unique())
    month_customers = set(m['Customer_ID'].unique())

    returning = month_customers & prior
    new_custs = month_customers - prior

    visit_days = m.groupby('Customer_ID')['Date'].apply(lambda s: s.dt.normalize().nunique())
    repeat_visitors = set(visit_days[visit_days >= 2].index)
    one_timers = set(visit_days[visit_days == 1].index)

    def rev_of(ids):
        return money(m[m['Customer_ID'].isin(ids)]['Total'].sum())

    def pct(n, d):
        return round(n / d * 100, 1) if d else 0

    total_cust = len(month_customers)
    freq = visit_days.value_counts().sort_index()

    retention = {
        'scope': scope,
        'totalCustomers': total_cust,
        'returningCustomers': len(returning),
        'returningPct': pct(len(returning), total_cust),
        'newCustomers': len(new_custs),
        'newPct': pct(len(new_custs), total_cust),
        'repeatVisitors': len(repeat_visitors),
        'repeatVisitorsPct': pct(len(repeat_visitors), total_cust),
        'oneTimeVisitors': len(one_timers),
        'oneTimeVisitorsPct': pct(len(one_timers), total_cust),
        'revenueReturning': rev_of(returning),
        'revenueReturningPct': pct(rev_of(returning), revenue),
        'revenueNew': rev_of(new_custs),
        'revenueNewPct': pct(rev_of(new_custs), revenue),
        'revenueRepeatVisitors': rev_of(repeat_visitors),
        'revenueRepeatVisitorsPct': pct(rev_of(repeat_visitors), revenue),
        'revenueOneTime': rev_of(one_timers),
        'revenueOneTimePct': pct(rev_of(one_timers), revenue),
        'visitFrequency': [
            {'days': int(d), 'customers': int(c), 'pct': pct(int(c), total_cust)}
            for d, c in freq.items()
        ],
    }

    return {
        'shop': str(m['Shop'].iloc[0]),
        'month': str(month),
        'monthLabel': period.strftime('%B %Y'),
        'spend': spend,
        'byProduct': by_product,
        'fastMovers': fast_movers,
        'dailyTrend': daily_trend,
        'bestCustomers': best_customers,
        'retention': retention,
    }


@app.route('/api/location-metrics')
def location_metrics():
    """Per-location, per-month metrics for the Shop Performance tab."""
    shop  = request.args.get('shop')
    month = request.args.get('month')
    scope = request.args.get('scope', 'shop')
    scope = scope if scope in ('shop', 'all') else 'shop'
    try:
        df = get_customer_data()

        shops = sorted(
            s for s in df['Shop'].dropna().astype(str).str.strip().unique() if s
        )
        if not shop or shop == 'all':
            return jsonify({'shops': shops, 'months': _location_month_options(df),
                            'metrics': None})

        months = _location_month_options(df, shop)
        if not month or month == 'all':
            month = months[0]['value'] if months else None
        if not month:
            return jsonify({'shops': shops, 'months': [], 'metrics': None})

        metrics = _calculate_location_metrics(df, shop, month, scope=scope)
        return jsonify({'shops': shops, 'months': months, 'metrics': metrics})
    except Exception as e:
        import traceback
        print(f"[ERROR] /api/location-metrics: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500




# ── Excel export of location metrics ─────────────────────────────────────────

_XL_HEAD = None  # lazily built styles, so openpyxl is only imported on demand


def _xl_styles():
    from openpyxl.styles import Font, PatternFill, Alignment
    return {
        'title': Font(bold=True, size=13, color='FFFFFF'),
        'title_fill': PatternFill('solid', fgColor='4F46E5'),
        'head': Font(bold=True, color='FFFFFF'),
        'head_fill': PatternFill('solid', fgColor='334155'),
        'sect': Font(bold=True, color='0F172A'),
        'sect_fill': PatternFill('solid', fgColor='E2E8F0'),
        'center': Alignment(horizontal='center'),
    }


def _xl_write(ws, rows, st, widths=None):
    """Write [(kind, cells)] rows where kind is title|head|sect|data."""
    for cells in rows:
        kind, values = cells[0], cells[1]
        r = ws.max_row + 1 if ws.max_row > 1 or ws['A1'].value is not None else 1
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=i, value=v)
            if kind == 'title':
                c.font, c.fill = st['title'], st['title_fill']
            elif kind == 'head':
                c.font, c.fill = st['head'], st['head_fill']
            elif kind == 'sect':
                c.font, c.fill = st['sect'], st['sect_fill']
    if widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_for_location(wb, m, first=False):
    """Write one location's four metric groups, mirroring the source workbook."""
    st = _xl_styles()
    tag = f"{m['shop']} {m['month']}"
    s, t = m['spend'], m['retention']
    scope_note = ("this shop's own history" if t['scope'] == 'shop'
                  else 'first visit at any shop')

    ws = wb.active if first else wb.create_sheet()
    ws.title = (f"{m['shop']} - Spend")[:31]
    _xl_write(ws, [
        ('title', [f"{tag} | Spend Metrics"]),
        ('sect', ['OVERALL SUMMARY']),
        ('data', ['Total Transactions', s['totalTransactions']]),
        ('data', ['Unique Customers', s['uniqueCustomers']]),
        ('data', ['Total Revenue (Ksh)', s['totalRevenue']]),
        ('sect', ['AVG SPEND / TRANSACTION']),
        ('data', ['Avg Spend per Transaction', s['avgPerTransaction']]),
        ('data', ['Min Spend (single txn)', s['minTransaction']]),
        ('data', ['Max Spend (single txn)', s['maxTransaction']]),
        ('sect', ['AVG SPEND / CUSTOMER']),
        ('data', ['Avg Spend per Customer', s['avgPerCustomer']]),
        ('data', ['Median Spend per Customer', s['medianPerCustomer']]),
        ('data', ['Highest Customer Spend', s['highestCustomerSpend']]),
        ('data', ['Lowest Customer Spend', s['lowestCustomerSpend']]),
        ('data', []),
        ('sect', ['AVG SPEND PER TRANSACTION BY PRODUCT']),
        ('head', ['Product', 'Units Sold', 'Transactions', 'Revenue (Ksh)',
                  'Avg / Transaction (Ksh)']),
    ], st, widths=[34, 14, 14, 18, 22])
    for p in m['byProduct']:
        _xl_write(ws, [('data', [p['product'], p['units'], p['transactions'],
                                 p['revenue'], p['avgPerTransaction']])], st)

    ws = wb.create_sheet((f"{m['shop']} - Fast Moving")[:31])
    _xl_write(ws, [
        ('title', [f"{tag} | Top 10 Fast Moving Products"]),
        ('head', ['Rank', 'Product', 'Units Sold', '% of All Units Sold']),
    ], st, widths=[8, 30, 14, 22])
    for f in m['fastMovers']:
        _xl_write(ws, [('data', [f['rank'], f['product'], f['units'],
                                 f['pctOfUnits'] / 100])], st)
    for r in range(3, 3 + len(m['fastMovers'])):
        ws.cell(row=r, column=4).number_format = '0.0%'

    dt = m['dailyTrend']
    if dt['series']:
        _xl_write(ws, [('data', []),
                       ('sect', [f"DAILY SALES TREND - TOP 10 ({m['monthLabel']})"]),
                       ('head', ['Day'] + [x['product'] for x in dt['series']])], st)
        for i, d in enumerate(dt['days']):
            _xl_write(ws, [('data', [d] + [x['data'][i] for x in dt['series']])], st)

    ws = wb.create_sheet((f"{m['shop']} - Best Custs")[:31])
    _xl_write(ws, [
        ('title', [f"{tag} | Best Customer Tracker"]),
        ('head', ['Rank', 'Name', 'Phone', 'Gender', 'Total Spend (Ksh)',
                  'Transactions', 'Visit Days', 'Avg/Transaction (Ksh)',
                  'Last Visit', 'Products Bought']),
    ], st, widths=[7, 18, 15, 10, 18, 14, 12, 22, 14, 60])
    for b in m['bestCustomers']:
        _xl_write(ws, [('data', [b['rank'], b['name'], b['phone'], b['gender'],
                                 b['totalSpend'], b['transactions'], b['visitDays'],
                                 b['avgPerTransaction'], b['lastVisit'],
                                 ', '.join(b['products'])])], st)

    ws = wb.create_sheet((f"{m['shop']} - Retention")[:31])
    _xl_write(ws, [
        ('title', [f"{tag} | Repeat Customers & Retention"]),
        ('head', ['Metric', 'Value', '% / Note']),
        ('sect', [f'CUSTOMER BASE (based on {scope_note})']),
        ('data', ['Total Unique Customers', t['totalCustomers'], '100%']),
        ('data', ['Returning Customers', t['returningCustomers'], f"{t['returningPct']}%"]),
        ('data', ['New Customers', t['newCustomers'], f"{t['newPct']}%"]),
        ('sect', [f"{m['monthLabel'].upper()} BEHAVIOUR"]),
        ('data', ['Repeat Visitors (2+ days)', t['repeatVisitors'], f"{t['repeatVisitorsPct']}%"]),
        ('data', ['One-Time Visitors (1 day)', t['oneTimeVisitors'], f"{t['oneTimeVisitorsPct']}%"]),
        ('sect', ['REVENUE SPLIT']),
        ('data', ['Revenue from Returning Custs', t['revenueReturning'], f"{t['revenueReturningPct']}% of total"]),
        ('data', ['Revenue from New Customers', t['revenueNew'], f"{t['revenueNewPct']}% of total"]),
        ('data', ['Revenue from Repeat Visitors', t['revenueRepeatVisitors'], f"{t['revenueRepeatVisitorsPct']}% of total"]),
        ('data', ['Revenue from One-Time Visitors', t['revenueOneTime'], f"{t['revenueOneTimePct']}% of total"]),
        ('data', []),
        ('sect', ['VISIT FREQUENCY DISTRIBUTION']),
        ('head', ['Visit Days in Month', 'Customers', '% of Customers']),
    ], st, widths=[34, 18, 20])
    for f in t['visitFrequency']:
        _xl_write(ws, [('data', [f"{f['days']} day{'' if f['days'] == 1 else 's'}",
                                 f['customers'], f"{f['pct']}%"])], st)


@app.route('/api/export/location-metrics')
def export_location_metrics():
    """Excel export: one location, or every location, for a month.

    Mirrors the layout of the source metrics workbook so the output can be
    dropped straight in beside the hand-built reports it replaces.
    """
    from flask import send_file
    from openpyxl import Workbook

    shop  = request.args.get('shop', 'all')
    month = request.args.get('month')
    scope = request.args.get('scope', 'shop')
    scope = scope if scope in ('shop', 'all') else 'shop'

    try:
        df = get_customer_data()
        months = _location_month_options(df, None if shop == 'all' else shop)
        if not month or month == 'all':
            month = months[0]['value'] if months else None
        if not month:
            return jsonify({'error': 'No data available to export'}), 400

        wb = Workbook()

        if shop and shop != 'all':
            m = _calculate_location_metrics(df, shop, month, scope=scope)
            if not m:
                return jsonify({'error': f'No data for {shop} in {month}'}), 400
            _sheet_for_location(wb, m, first=True)
            fname = f"{shop}_{month}_metrics.xlsx".replace(' ', '_')
        else:
            # All locations: a comparison sheet first, then each shop's detail.
            shops = sorted(s for s in df['Shop'].dropna().astype(str).str.strip().unique() if s)
            computed = []
            for sname in shops:
                m = _calculate_location_metrics(df, sname, month, scope=scope)
                if m:
                    computed.append(m)
            if not computed:
                return jsonify({'error': f'No data for any location in {month}'}), 400

            st = _xl_styles()
            ws = wb.active
            ws.title = 'All Locations'
            _xl_write(ws, [
                ('title', [f"All Locations | {computed[0]['monthLabel']}"]),
                ('head', ['Location', 'Transactions', 'Unique Customers',
                          'Revenue (Ksh)', 'Avg/Transaction', 'Avg/Customer',
                          'Returning', 'Returning %', 'New', 'New %',
                          'Repeat Visitors', 'Repeat %', 'One-Time', 'One-Time %',
                          'Top Product']),
            ], st, widths=[18, 14, 18, 18, 16, 16, 12, 12, 10, 10, 16, 12, 12, 12, 24])
            for m in sorted(computed, key=lambda x: -x['spend']['totalRevenue']):
                s, t = m['spend'], m['retention']
                _xl_write(ws, [('data', [
                    m['shop'], s['totalTransactions'], s['uniqueCustomers'],
                    s['totalRevenue'], s['avgPerTransaction'], s['avgPerCustomer'],
                    t['returningCustomers'], f"{t['returningPct']}%",
                    t['newCustomers'], f"{t['newPct']}%",
                    t['repeatVisitors'], f"{t['repeatVisitorsPct']}%",
                    t['oneTimeVisitors'], f"{t['oneTimeVisitorsPct']}%",
                    m['fastMovers'][0]['product'] if m['fastMovers'] else '',
                ])], st)
            ws.freeze_panes = 'A3'

            for m in computed:
                _sheet_for_location(wb, m)
            fname = f"all_locations_{month}_metrics.xlsx"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf, as_attachment=True, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        import traceback
        print(f"[ERROR] /api/export/location-metrics: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5002))
    app.run(host="0.0.0.0", port=port)
else:
    # When launched by gunicorn, pre-warm the cache in a background thread.
    # This means by the time the first browser request arrives, all the heavy
    # computation is already done and results are served instantly from cache.
    def _prewarm():
        try:
            print("[INFO] Pre-warming data cache in background thread...")
            with app.app_context():
                from flask import Request
                with app.test_request_context('/'):
                        # Directly call the internal compute path
                    df = get_customer_data()
                    if df is not None and not df.empty:
                        _compute_all_results(df)
                        print("[INFO] Pre-warm complete. Data is ready.")
        except Exception as e:
            print(f"[WARNING] Pre-warm failed (non-fatal): {e}")

    threading.Thread(target=_prewarm, daemon=True).start()